"""管理端 API：概览统计 / 用户管理 / 用量统计 / 模型配置 / 系统设置 / 导入导出记录。"""

import io
import json
import os
import re

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .auth import (
    ROLE_SUPER_ADMIN,
    ROLE_SYSTEM_ADMIN,
    compute_age,
    hash_password,
    require_admin,
    require_model_admin,
    require_settings_admin,
    require_super_admin,
)
from .config import settings
from .db import DEFAULT_MODEL_PROVIDERS, execute, fetch_all, fetch_one, now_ms, transaction
from .schemas import (
    DimTableCreate,
    DimTableOut,
    DimTableUpdate,
    DimValueCreate,
    DimValueList,
    DimValueOut,
    DimValueUpdate,
    ModelPayload,
    OperationLogList,
    ResetPasswordRequest,
    SettingLogList,
    SettingsPayload,
    UserUpdate,
)

router = APIRouter(prefix="/api/admin", tags=["admin"], dependencies=[Depends(require_admin)])

_DAY_MS = 24 * 3600 * 1000

_PERIODS: dict[str, int] = {
    "day": 1 * _DAY_MS,
    "week": 7 * _DAY_MS,
    "month": 30 * _DAY_MS,
    "year": 365 * _DAY_MS,
}


@router.get("/stats")
def stats():
    now = now_ms()
    users = fetch_one("SELECT COUNT(*) AS n FROM users")["n"]
    active_today = fetch_one(
        "SELECT COUNT(DISTINCT user_id) AS n FROM token_usage WHERE created_at >= ?",
        (now - _DAY_MS,),
    )["n"]
    active_7d = fetch_one(
        "SELECT COUNT(DISTINCT user_id) AS n FROM token_usage WHERE created_at >= ?",
        (now - 7 * _DAY_MS,),
    )["n"]
    total_tokens = fetch_one(
        "SELECT COALESCE(SUM(total_tokens), 0) AS n FROM token_usage"
    )["n"]
    today_tokens = fetch_one(
        "SELECT COALESCE(SUM(total_tokens), 0) AS n FROM token_usage WHERE created_at >= ?",
        (now - _DAY_MS,),
    )["n"]
    requests = fetch_one("SELECT COUNT(*) AS n FROM token_usage")["n"]
    return {
        "users": users,
        "active_today": active_today,
        "active_7d": active_7d,
        "total_tokens": total_tokens,
        "today_tokens": today_tokens,
        "requests": requests,
    }


@router.get("/overview")
def overview():
    """管理台概览：核心指标 + 多维图表数据（趋势/时段/新增/排行/画像）。"""
    now = now_ms()
    start30 = now - 30 * _DAY_MS

    users = fetch_one("SELECT COUNT(*) AS n FROM users")["n"]
    active_today = fetch_one(
        "SELECT COUNT(DISTINCT user_id) AS n FROM token_usage WHERE created_at >= ?",
        (now - _DAY_MS,),
    )["n"]
    active_7d = fetch_one(
        "SELECT COUNT(DISTINCT user_id) AS n FROM token_usage WHERE created_at >= ?",
        (now - 7 * _DAY_MS,),
    )["n"]
    total_tokens = fetch_one(
        "SELECT COALESCE(SUM(total_tokens), 0) AS n FROM token_usage"
    )["n"]
    today_tokens = fetch_one(
        "SELECT COALESCE(SUM(total_tokens), 0) AS n FROM token_usage WHERE created_at >= ?",
        (now - _DAY_MS,),
    )["n"]
    requests = fetch_one("SELECT COUNT(*) AS n FROM token_usage")["n"]

    daily = fetch_all(
        """
        SELECT (created_at / 86400000) AS day, COUNT(*) AS requests,
               COALESCE(SUM(total_tokens), 0) AS total
        FROM token_usage WHERE created_at >= ? GROUP BY day ORDER BY day
        """,
        (start30,),
    )

    hourly = fetch_all(
        """
        SELECT (created_at / 3600000) % 24 AS hour, COUNT(*) AS requests,
               COALESCE(SUM(total_tokens), 0) AS total
        FROM token_usage WHERE created_at >= ? GROUP BY hour ORDER BY hour
        """,
        (start30,),
    )

    new_users = fetch_all(
        """
        SELECT (created_at / 86400000) AS day, COUNT(*) AS n
        FROM users WHERE created_at >= ? GROUP BY day ORDER BY day
        """,
        (start30,),
    )

    by_model = fetch_all(
        """
        SELECT t.model_key AS model_key,
               COALESCE(m.name, t.model_key) AS name,
               COALESCE(m.name_en, '') AS name_en,
               COALESCE(m.provider, '') AS provider,
               COUNT(*) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total
        FROM token_usage t LEFT JOIN models m ON m.id = COALESCE(t.model_id,
             (SELECT m2.id FROM models m2 WHERE m2.model_key = t.model_key
              ORDER BY m2.sort_order, m2.id LIMIT 1))
        WHERE t.created_at >= ?
        GROUP BY COALESCE(t.model_id,
             (SELECT m3.id FROM models m3 WHERE m3.model_key = t.model_key
              ORDER BY m3.sort_order, m3.id LIMIT 1),
             'key:' || t.model_key)
        ORDER BY total DESC LIMIT 8
        """,
        (start30,),
    )

    by_provider = fetch_all(
        """
        SELECT COALESCE(m.provider, '') AS provider,
               COUNT(*) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total
        FROM token_usage t LEFT JOIN models m ON m.id = COALESCE(t.model_id,
             (SELECT m2.id FROM models m2 WHERE m2.model_key = t.model_key
              ORDER BY m2.sort_order, m2.id LIMIT 1))
        WHERE t.created_at >= ? AND COALESCE(m.provider, '') != ''
        GROUP BY COALESCE(m.provider, '')
        ORDER BY total DESC
        """,
        (start30,),
    )

    top_users = fetch_all(
        """
        SELECT u.username, u.avatar, u.province, u.city,
               COALESCE(SUM(t.total_tokens), 0) AS total,
               COUNT(t.id) AS requests
        FROM users u JOIN token_usage t ON t.user_id = u.id
        WHERE t.created_at >= ?
        GROUP BY u.id ORDER BY total DESC LIMIT 8
        """,
        (start30,),
    )

    top_provinces = fetch_all(
        """
        SELECT u.province,
               COUNT(DISTINCT t.user_id) AS active_users,
               COUNT(t.id) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total_tokens
        FROM users u JOIN token_usage t ON t.user_id = u.id
        WHERE u.province != '' AND t.created_at >= ?
        GROUP BY u.province ORDER BY active_users DESC, total_tokens DESC LIMIT 8
        """,
        (start30,),
    )

    recent_users = fetch_all(
        """
        SELECT username, avatar, province, city, district, created_at
        FROM users ORDER BY created_at DESC LIMIT 8
        """
    )

    age_dist, gender_dist = _aggregate_demographics()

    return {
        "stats": {
            "users": users,
            "active_today": active_today,
            "active_7d": active_7d,
            "total_tokens": total_tokens,
            "today_tokens": today_tokens,
            "requests": requests,
        },
        "daily": [dict(r) for r in daily],
        "hourly": [dict(r) for r in hourly],
        "new_users": [dict(r) for r in new_users],
        "by_model": [dict(r) for r in by_model],
        "by_provider": [dict(r) for r in by_provider],
        "top_users": [dict(r) for r in top_users],
        "top_provinces": [dict(r) for r in top_provinces],
        "recent_users": [dict(r) for r in recent_users],
        "age_dist": age_dist,
        "gender_dist": gender_dist,
    }


@router.get("/users")
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    search: str = "",
    username: str = "",
    gender: str = "",
    role: str = "",
    is_active: str = "",
    sort: str = "",
    order: str = "desc",
):
    offset = (page - 1) * page_size
    clauses: list[str] = []
    params: list = []
    if search:
        clauses.append("u.username LIKE ?")
        params.append(f"%{search}%")
    if username:
        clauses.append("u.username LIKE ?")
        params.append(f"%{username}%")
    if gender:
        genders = [g.strip() for g in gender.split(",") if g.strip()]
        if genders:
            clauses.append("u.gender IN (" + ",".join("?" for _ in genders) + ")")
            params.extend(genders)
    if role:
        roles = [r.strip() for r in role.split(",") if r.strip()]
        if roles:
            clauses.append("u.role IN (" + ",".join("?" for _ in roles) + ")")
            params.extend(roles)
    if is_active:
        clauses.append("u.is_active = ?")
        params.append(1 if is_active.strip().lower() == "true" else 0)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    # 排序（字段白名单，防 SQL 注入）
    sort_columns = {
        "username": "u.username",
        "role": "u.role",
        "is_active": "u.is_active",
        "gender": "u.gender",
        "birthday": "u.birthday",
        "created_at": "u.created_at",
        "updated_at": "u.updated_at",
        "last_seen_at": "u.last_seen_at",
        "logins": "logins",
        "total_tokens": "total_tokens",
    }
    order_by = "u.created_at DESC"
    if sort in sort_columns:
        direction = "ASC" if order.strip().lower() == "asc" else "DESC"
        order_by = f"{sort_columns[sort]} {direction}"
    total = fetch_one(f"SELECT COUNT(*) AS n FROM users u {where}", params)["n"]
    rows = fetch_all(
        f"""
        SELECT u.id, u.username, u.role, u.is_active, u.created_at, u.last_seen_at,
               u.province, u.city, u.district, u.age, u.birthday, u.gender,
               u.updated_at, u.updated_by,
               COALESCE((SELECT COUNT(*) FROM login_logs l WHERE l.user_id = u.id AND l.success = 1), 0) AS logins,
               COALESCE((SELECT SUM(t.total_tokens) FROM token_usage t WHERE t.user_id = u.id), 0) AS total_tokens
        FROM users u {where} ORDER BY {order_by}
        LIMIT ? OFFSET ?
        """,
        (*params, page_size, offset),
    )
    return {"items": [dict(r) for r in rows], "total": total, "page": page, "pageSize": page_size}


# 用户可导出字段顺序（与 xlsx 表头一致）
_USER_FIELDS = [
    "username", "role", "is_active", "birthday", "gender", "province", "city", "district",
    "logins", "total_tokens", "last_seen_at", "created_at", "updated_at", "updated_by",
]


# 注意：必须注册在 /users/{user_id} 之前，否则 "export" 会被 {user_id} 路由拦截（int 解析失败 → 422）
@router.get("/users/export")
def export_users(admin: dict = Depends(require_super_admin)):
    """导出全部用户为 xlsx。导出不存产物：仅写行为记录。"""
    username = admin.get("username") or "admin"
    rows = fetch_all(
        """
        SELECT u.username, u.role, u.is_active, u.birthday, u.gender, u.province, u.city, u.district,
               COALESCE((SELECT COUNT(*) FROM login_logs l WHERE l.user_id = u.id AND l.success = 1), 0) AS logins,
               COALESCE((SELECT SUM(t.total_tokens) FROM token_usage t WHERE t.user_id = u.id), 0) AS total_tokens,
               u.last_seen_at, u.created_at, u.updated_at, u.updated_by
        FROM users u ORDER BY u.id
        """
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "users"
    ws.append(_USER_FIELDS)
    for r in rows:
        ws.append(["" if r[f] is None else r[f] for f in _USER_FIELDS])
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    filename = f"users_{now_ms()}.xlsx"
    _write_transfer_meta(
        "export", username, filename, len(data), _XLSX_MIME,
        remark="用户数据导出", status="success",
    )
    return StreamingResponse(
        iter([data]),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/users/{user_id}")
def get_user(user_id: int):
    row = fetch_one(
        """
        SELECT u.id, u.username, u.role, u.is_active, u.created_at, u.last_seen_at,
               u.province, u.city, u.district, u.age, u.birthday, u.gender,
               u.updated_at, u.updated_by,
               COALESCE((SELECT COUNT(*) FROM login_logs l WHERE l.user_id = u.id AND l.success = 1), 0) AS logins,
               COALESCE((SELECT SUM(t.total_tokens) FROM token_usage t WHERE t.user_id = u.id), 0) AS total_tokens
        FROM users u WHERE u.id = ?
        """,
        (user_id,),
    )
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    return dict(row)


_ROLE_LABELS = {
    "super_admin": ("超级管理员", "Super Admin"),
    "system_admin": ("系统管理员", "System Admin"),
    "model_admin": ("模型管理员", "Model Admin"),
    "user": ("普通用户", "User"),
    "subscriber": ("订阅用户", "Subscriber"),
}

_GENDER_LABELS = {
    "male": ("男", "Male"),
    "female": ("女", "Female"),
    "other": ("其他", "Other"),
}


def _user_diff_logs(old: dict, body) -> list[tuple[str, str]]:
    """用户字段级变更日志：逐字段比较旧值→新值，返回 (中文, English) 列表。"""
    old = dict(old)  # 兼容 sqlite3.Row（fetch_one 返回）
    msgs: list[tuple[str, str]] = []
    if body.is_active is not None and bool(old["is_active"]) != bool(body.is_active):
        o = "启用" if bool(old["is_active"]) else "禁用"
        n = "启用" if body.is_active else "禁用"
        oe = "enabled" if bool(old["is_active"]) else "disabled"
        ne = "enabled" if body.is_active else "disabled"
        msgs.append((f'状态从「{o}」改成「{n}」', f'Status changed from "{oe}" to "{ne}"'))
    if body.role is not None and old["role"] != body.role:
        o_zh, o_en = _ROLE_LABELS.get(old["role"], (old["role"], old["role"]))
        n_zh, n_en = _ROLE_LABELS.get(body.role, (body.role, body.role))
        msgs.append((f'角色从「{o_zh}」改成「{n_zh}」', f'Role changed from "{o_en}" to "{n_en}"'))
    if body.gender is not None and (old.get("gender") or "") != body.gender:
        o_zh, o_en = _GENDER_LABELS.get(old.get("gender") or "", ("未设置", "Not set"))
        n_zh, n_en = _GENDER_LABELS.get(body.gender, (body.gender, body.gender))
        msgs.append((f'性别从「{o_zh}」改成「{n_zh}」', f'Gender changed from "{o_en}" to "{n_en}"'))
    if body.birthday is not None and (old.get("birthday") or "") != body.birthday:
        old_b = old.get("birthday") or ""
        new_b = body.birthday or ""
        msgs.append((f'生日从「{old_b or "未设置"}」改成「{new_b or "未设置"}」',
                     f'Birthday changed from "{old_b or "Not set"}" to "{new_b or "Not set"}"'))
    if body.province is not None and (old.get("province") or "") != body.province:
        msgs.append((f'省份从「{old.get("province") or ""}」改成「{body.province}」',
                     f'Province changed from "{old.get("province") or ""}" to "{body.province}"'))
    if body.city is not None and (old.get("city") or "") != body.city:
        msgs.append((f'城市从「{old.get("city") or ""}」改成「{body.city}」',
                     f'City changed from "{old.get("city") or ""}" to "{body.city}"'))
    if body.district is not None and (old.get("district") or "") != body.district:
        msgs.append((f'区县从「{old.get("district") or ""}」改成「{body.district}」',
                     f'District changed from "{old.get("district") or ""}" to "{body.district}"'))
    return msgs


@router.patch("/users/{user_id}")
def update_user(user_id: int, body: UserUpdate, admin: dict = Depends(require_super_admin)):
    row = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    sets, params = [], []
    if body.is_active is not None:
        sets.append("is_active = ?")
        params.append(1 if body.is_active else 0)
    if body.role is not None:
        sets.append("role = ?")
        params.append(body.role)
    if body.province is not None:
        sets.append("province = ?")
        params.append(body.province)
    if body.city is not None:
        sets.append("city = ?")
        params.append(body.city)
    if body.district is not None:
        sets.append("district = ?")
        params.append(body.district)
    if body.age is not None:
        sets.append("age = ?")
        params.append(body.age)
    if body.birthday is not None:
        sets.append("birthday = ?")
        params.append(body.birthday)
        sets.append("age = ?")
        params.append(compute_age(body.birthday) if body.birthday else None)
    if body.gender is not None:
        sets.append("gender = ?")
        params.append(body.gender)
    if not sets:
        return {"ok": True}
    sets.append("updated_at = ?")
    params.append(now_ms())
    sets.append("updated_by = ?")
    params.append(admin["username"])
    params.append(user_id)
    execute(f"UPDATE users SET {', '.join(sets)} WHERE id = ?", params)
    # 字段级变更日志（旧值 → 新值）
    for zh, en in _user_diff_logs(row, body):
        _log_operation("user", user_id, admin, zh, en)
    return {"ok": True}


@router.delete("/users/{user_id}")
def delete_user(user_id: int, admin: dict = Depends(require_super_admin)):
    """删除用户：仅超级管理员可执行，且不能删除自己、只能删除未启用的用户。"""
    if user_id == admin["id"]:
        raise HTTPException(status_code=400, detail="不能删除自己")
    row = fetch_one("SELECT * FROM users WHERE id = ?", (user_id,))
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    if row["is_active"]:
        raise HTTPException(status_code=400, detail="启用状态的用户不能删除，请先禁用")
    # 清理关联数据（login_logs/token_usage 无外键，需手动删；sessions/messages 随 users 级联删除）
    transaction([
        ("DELETE FROM login_logs WHERE user_id = ?", (user_id,)),
        ("DELETE FROM token_usage WHERE user_id = ?", (user_id,)),
        ("DELETE FROM users WHERE id = ?", (user_id,)),
    ])
    _log_operation("user", user_id, admin, f"删除用户「{row['username']}」", f'Deleted user "{row["username"]}"')
    return {"ok": True}


@router.post("/users/{user_id}/reset-password")
def reset_password(user_id: int, body: ResetPasswordRequest, admin: dict = Depends(require_super_admin)):
    row = fetch_one("SELECT id, username FROM users WHERE id = ?", (user_id,))
    if not row:
        raise HTTPException(status_code=404, detail="用户不存在")
    execute(
        "UPDATE users SET password_hash = ?, updated_at = ?, updated_by = ? WHERE id = ?",
        (hash_password(body.password), now_ms(), admin["username"], user_id),
    )
    _log_operation(
        "user", user_id, admin,
        f"重置用户「{row['username']}」的密码",
        f'Reset password for user "{row["username"]}"',
    )
    return {"ok": True}


@router.get("/usage")
def usage():
    now = now_ms()
    by_user = fetch_all(
        """
        SELECT u.username, COALESCE(SUM(t.total_tokens), 0) AS total,
               COALESCE(SUM(t.prompt_tokens), 0) AS prompt,
               COALESCE(SUM(t.completion_tokens), 0) AS completion,
               COUNT(t.id) AS requests
        FROM users u LEFT JOIN token_usage t ON t.user_id = u.id
        GROUP BY u.id ORDER BY total DESC
        """
    )
    by_model = fetch_all(
        """
        SELECT t.model_key AS model_key,
               COALESCE(m.name, t.model_key) AS name,
               COALESCE(m.name_en, '') AS name_en,
               COALESCE(m.provider, '') AS provider,
               COUNT(*) AS requests,
               COALESCE(SUM(t.prompt_tokens), 0) AS prompt,
               COALESCE(SUM(t.completion_tokens), 0) AS completion,
               COALESCE(SUM(t.total_tokens), 0) AS total
        FROM token_usage t LEFT JOIN models m ON m.id = COALESCE(t.model_id,
             (SELECT m2.id FROM models m2 WHERE m2.model_key = t.model_key
              ORDER BY m2.sort_order, m2.id LIMIT 1))
        GROUP BY COALESCE(t.model_id,
             (SELECT m3.id FROM models m3 WHERE m3.model_key = t.model_key
              ORDER BY m3.sort_order, m3.id LIMIT 1),
             'key:' || t.model_key)
        ORDER BY total DESC
        """
    )
    daily = fetch_all(
        """
        SELECT (created_at / 86400000) AS day, COUNT(*) AS requests,
               COALESCE(SUM(total_tokens), 0) AS total
        FROM token_usage WHERE created_at >= ? GROUP BY day ORDER BY day
        """,
        (now - 30 * _DAY_MS,),
    )
    by_provider = fetch_all(
        """
        SELECT COALESCE(m.provider, '') AS provider,
               COUNT(*) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total
        FROM token_usage t LEFT JOIN models m ON m.id = COALESCE(t.model_id,
             (SELECT m2.id FROM models m2 WHERE m2.model_key = t.model_key
              ORDER BY m2.sort_order, m2.id LIMIT 1))
        WHERE COALESCE(m.provider, '') != ''
        GROUP BY COALESCE(m.provider, '')
        ORDER BY total DESC
        """
    )
    by_city = fetch_all(
        """
        SELECT u.province AS province, u.city AS city,
               COUNT(t.id) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total
        FROM users u JOIN token_usage t ON t.user_id = u.id
        WHERE u.city != ''
        GROUP BY u.province, u.city
        ORDER BY total DESC LIMIT 60
        """
    )
    age_dist, gender_dist = _aggregate_demographics()
    return {
        "by_user": [dict(r) for r in by_user],
        "by_model": [dict(r) for r in by_model],
        "by_provider": [dict(r) for r in by_provider],
        "by_city": [dict(r) for r in by_city],
        "daily": [dict(r) for r in daily],
        "age_dist": age_dist,
        "gender_dist": gender_dist,
    }


_AGE_ORDER = ["0-17", "18-24", "25-34", "35-44", "45-54", "55-64", "65+", "unknown"]
_GENDER_ORDER = ["male", "female", "other", "unknown"]


def _aggregate_demographics() -> tuple[list[dict], list[dict]]:
    """按行业标准年龄段与性别聚合：
    - 年龄分布统计 Token 消耗量（按年龄分组的 token_usage 总量）
    - 性别分布统计用户人数；未填写(生日/性别)统一归为 unknown。"""
    rows = fetch_all(
        """
        SELECT u.age, u.birthday, u.gender, COALESCE(SUM(t.total_tokens), 0) AS tokens
        FROM users u LEFT JOIN token_usage t ON t.user_id = u.id
        GROUP BY u.id
        """
    )
    age_map = {k: 0 for k in _AGE_ORDER}
    gender_map = {k: 0 for k in _GENDER_ORDER}
    for r in rows:
        age = compute_age(r["birthday"]) if r["birthday"] else r["age"]
        if age is None:
            age_key = "unknown"
        elif age < 18:
            age_key = "0-17"
        elif age <= 24:
            age_key = "18-24"
        elif age <= 34:
            age_key = "25-34"
        elif age <= 44:
            age_key = "35-44"
        elif age <= 54:
            age_key = "45-54"
        elif age <= 64:
            age_key = "55-64"
        else:
            age_key = "65+"
        age_map[age_key] += r["tokens"]
        g = (r["gender"] or "").strip().lower()
        gender_key = g if g in ("male", "female", "other") else "unknown"
        gender_map[gender_key] += 1
    age_dist = [{"key": k, "count": age_map[k]} for k in _AGE_ORDER]
    gender_dist = [{"key": k, "count": gender_map[k]} for k in _GENDER_ORDER]
    return age_dist, gender_dist


@router.get("/region-stats")
def region_stats(period: str = "month"):
    """按统计周期聚合省份热度指标与省/市/区用户分布。

    - provinces: 每个省份的 新增用户数 / 活跃用户数 / 对话次数 / Token 消耗量（仅统计周期内）
    - regions:   省/市/区粒度用户分布 + 活跃 TOP3（requests/tokens 限定统计周期）
    """
    if period not in _PERIODS:
        raise HTTPException(status_code=400, detail="period 仅支持 day/week/month/year")
    start = now_ms() - _PERIODS[period]

    prov_rows = fetch_all(
        """
        SELECT u.province,
               COUNT(DISTINCT CASE WHEN u.created_at >= ? THEN u.id END) AS new_users,
               COUNT(DISTINCT t.user_id) AS active_users,
               COUNT(t.id) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total_tokens
        FROM users u
        LEFT JOIN token_usage t ON t.user_id = u.id AND t.created_at >= ?
        WHERE u.province != ''
        GROUP BY u.province
        """,
        (start, start),
    )
    provinces = [
        {
            "province": r["province"],
            "new_users": r["new_users"],
            "active_users": r["active_users"],
            "requests": r["requests"],
            "total_tokens": r["total_tokens"],
        }
        for r in prov_rows
    ]

    rows = fetch_all(
        """
        SELECT u.username, u.avatar, u.province, u.city, u.district,
               COUNT(t.id) AS requests,
               COALESCE(SUM(t.total_tokens), 0) AS total_tokens
        FROM users u
        LEFT JOIN token_usage t ON t.user_id = u.id AND t.created_at >= ?
        WHERE u.province != '' OR u.city != '' OR u.district != ''
        GROUP BY u.id
        ORDER BY u.province, u.city, u.district
        """,
        (start,),
    )
    regions: dict[tuple[str, str, str], dict] = {}
    for r in rows:
        key = (r["province"], r["city"], r["district"])
        reg = regions.setdefault(
            key,
            {
                "province": r["province"],
                "city": r["city"],
                "district": r["district"],
                "count": 0,
                "top_users": [],
            },
        )
        reg["count"] += 1
        reg["top_users"].append(
            {
                "username": r["username"],
                "avatar": r["avatar"] if r["avatar"] else "",
                "requests": r["requests"],
                "total_tokens": r["total_tokens"],
            }
        )
    for reg in regions.values():
        reg["top_users"].sort(key=lambda u: (u["requests"], u["total_tokens"]), reverse=True)
        reg["top_users"] = reg["top_users"][:10]
    return {"period": period, "provinces": provinces, "regions": list(regions.values())}


@router.get("/models")
def list_models(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    search: str = "",
    enabled: str = "",
    free: str = "",
    vision: str = "",
    supports_search: str = "",
    name: str = "",
    model_key: str = "",
    provider: str = "",
    sort: str = "",
    order: str = "asc",
):
    offset = (page - 1) * page_size
    clauses: list[str] = []
    params: list = []
    if search:
        clauses.append("(model_key LIKE ? OR name LIKE ? OR provider LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    # 名称 / 模型标识独立模糊搜索（表头 input 筛选）
    if name:
        clauses.append("(name LIKE ? OR name_en LIKE ?)")
        params.extend([f"%{name}%", f"%{name}%"])
    if model_key:
        clauses.append("model_key LIKE ?")
        params.append(f"%{model_key}%")
    if enabled:
        clauses.append("enabled = ?")
        params.append(1 if enabled.strip().lower() == "true" else 0)
    if free:
        clauses.append("free = ?")
        params.append(1 if free.strip().lower() == "true" else 0)
    if vision:
        clauses.append("vision = ?")
        params.append(1 if vision.strip().lower() == "true" else 0)
    if supports_search:
        clauses.append("supports_search = ?")
        params.append(1 if supports_search.strip().lower() == "true" else 0)
    if provider:
        providers = [p.strip() for p in provider.split(",") if p.strip()]
        if providers:
            clauses.append("provider IN (" + ",".join("?" for _ in providers) + ")")
            params.extend(providers)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    # 排序（字段白名单，防注入）
    sort_columns = {
        "model_key": "model_key",
        "name": "name",
        "provider": "provider",
        "sort_order": "sort_order",
        "free": "free",
        "enabled": "enabled",
        "vision": "vision",
        "supports_search": "supports_search",
    }
    order_by = "sort_order, id"
    if sort in sort_columns:
        direction = "ASC" if order.strip().lower() == "asc" else "DESC"
        order_by = f"{sort_columns[sort]} {direction}, id"
    total = fetch_one(f"SELECT COUNT(*) AS n FROM models {where}", params)["n"]
    rows = fetch_all(
        f"SELECT * FROM models {where} ORDER BY {order_by} LIMIT ? OFFSET ?",
        (*params, page_size, offset),
    )
    return {"items": [dict(r) for r in rows], "total": total, "page": page, "pageSize": page_size}


# 模型可导入导出的字段顺序（与 xlsx 表头一致）
_MODEL_FIELDS = ["model_key", "name", "name_en", "provider", "free", "vision", "supports_search", "enabled", "sort_order", "is_default"]


def _model_rows() -> list[list]:
    rows = fetch_all("SELECT * FROM models ORDER BY sort_order, id")
    return [[r[f] if r[f] is not None else "" for f in _MODEL_FIELDS] for r in rows]


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _write_transfer_record(
    type_: str, username: str, filename: str, data: bytes, mime_type: str,
    remark: str = "", status: str = "success",
) -> tuple[str, int]:
    """把源文件落盘到 transfers 目录，并写入导入/导出记录。返回 (文件绝对路径, 记录 id)。"""
    os.makedirs(_TRANSFER_DIR, exist_ok=True)
    ts = now_ms()
    safe_name = f"{ts}_{filename}"
    path = os.path.join(_TRANSFER_DIR, safe_name)
    with open(path, "wb") as f:
        f.write(data)
    record_id = execute(
        "INSERT INTO transfer_records (type, username, filename, file_size, mime_type, file_path, remark, status, created_at)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (type_, username, filename, len(data), mime_type, path, remark, status, ts),
    )
    return path, record_id


def _write_transfer_meta(
    type_: str, username: str, filename: str, size: int, mime_type: str,
    remark: str = "", status: str = "success",
) -> int:
    """仅写入导入/导出行为记录，不落盘文件（导出不存产物场景）。记录页可据此重新生成下载。"""
    return execute(
        "INSERT INTO transfer_records (type, username, filename, file_size, mime_type, file_path, remark, status, created_at)"
        " VALUES (?, ?, ?, ?, ?, '', ?, ?, ?)",
        (type_, username, filename, size, mime_type, remark, status, now_ms()),
    )


def _build_annotated_file(raw: bytes, row_errors: dict, filename: str) -> tuple[str, int]:
    """基于导入源文件生成带「错误分析」列的 xlsx：表头红底白字，错误行单元格黄底。返回 (path, size)。"""
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    err_col = ws.max_column + 1
    header_cell = ws.cell(row=1, column=err_col, value="错误分析")
    header_cell.fill = PatternFill(fill_type="solid", start_color="C00000", end_color="C00000")
    header_cell.font = Font(color="FFFFFF", bold=True)
    yellow = PatternFill(fill_type="solid", start_color="FFEB9C", end_color="FFEB9C")
    for row_no, reason in sorted(row_errors.items()):
        cell = ws.cell(row=row_no, column=err_col, value=reason)
        cell.fill = yellow
    os.makedirs(_TRANSFER_DIR, exist_ok=True)
    path = os.path.join(_TRANSFER_DIR, f"{now_ms()}_annotated_{filename}")
    wb.save(path)
    return path, os.path.getsize(path)


@router.get("/models/export")
def export_models(admin: dict = Depends(require_model_admin)):
    """导出全部模型为 xlsx（Excel 表格）。导出不存产物：仅写行为记录，记录页可重新生成下载。"""
    username = admin.get("username") or "admin"
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "models"
        ws.append(_MODEL_FIELDS)
        for row in _model_rows():
            ws.append(row)
        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
    except Exception as err:
        _write_transfer_meta(
            "export", username, f"models_{now_ms()}.xlsx", 0, _XLSX_MIME,
            remark=f"模型数据导出失败：{err}", status="failed",
        )
        raise

    filename = f"models_{now_ms()}.xlsx"
    _write_transfer_meta(
        "export", username, filename, len(data), _XLSX_MIME,
        remark="模型数据导出", status="success",
    )
    return StreamingResponse(
        iter([data]),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _get_provider_ids() -> list[str]:
    """读取 model_provider 维表的启用值 code 列表（供模型导入模板下拉）。"""
    table = fetch_one("SELECT id FROM dim_tables WHERE code = 'model_provider'")
    if not table:
        return []
    rows = fetch_all(
        "SELECT code FROM dim_values WHERE table_id = ? AND enabled = 1 ORDER BY sort_order, id",
        (table["id"],),
    )
    return [r["code"] for r in rows]


def _known_provider_ids() -> list[str]:
    """provider 合法取值：维表优先，维表未配置时回退本地默认字典。"""
    return _get_provider_ids() or [code for code, _, _ in DEFAULT_MODEL_PROVIDERS]


@router.get("/models/template")
def download_model_template(request: Request, admin: dict = Depends(require_model_admin)):
    """下载模型导入模板：表头 + 示例行 + 枚举列下拉校验（provider / 布尔列）。
    按 Accept-Language 本地化：英文状态下载英文模板（示例值与校验文案）。"""
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "models"
    ws.append(_MODEL_FIELDS)
    # 示例行：与导入解析逻辑对齐（布尔列「是/否」或 yes/no，_parse_bool 均可解析）
    if is_en:
        ws.append(["example/model-key", "Model Name (Chinese)", "Example Model Name", "openai", "yes", "no", "yes", "yes", 100, "no"])
        prov_err, prov_err_title = "Please select a provider from the dropdown", "Invalid value"
        bool_list, bool_err, bool_err_title = '"yes,no"', 'Please choose "yes" or "no"', "Invalid value"
    else:
        ws.append(["example/model-key", "示例模型名称", "Example Model Name", "openai", "是", "否", "是", "是", 100, "否"])
        prov_err, prov_err_title = "请从下拉列表选择 provider", "非法值"
        bool_list, bool_err, bool_err_title = '"是,否"', "请选择「是」或「否」", "非法值"

    # provider 列（第 4 列 D）下拉：实时读取数据字典，写入隐藏 dict 表 + 命名范围
    # （维表变动后下次下载即生效；命名范围不受 Excel 字面量列表 255 字符限制）
    provider_ids = _get_provider_ids() or [code for code, _, _ in DEFAULT_MODEL_PROVIDERS]
    if provider_ids:
        dict_ws = wb.create_sheet("dict")
        for i, pid in enumerate(provider_ids, start=1):
            dict_ws.cell(row=i, column=1, value=pid)
        dict_ws.sheet_state = "hidden"
        wb.defined_names["ProviderList"] = DefinedName(
            "ProviderList", attr_text=f"'dict'!$A$1:$A${len(provider_ids)}"
        )
        dv_provider = DataValidation(
            type="list",
            formula1="=ProviderList",
            allow_blank=True,
            showDropDown=False,
        )
        dv_provider.error = prov_err
        dv_provider.errorTitle = prov_err_title
        ws.add_data_validation(dv_provider)
        dv_provider.add("D2:D1000")

    # 布尔列下拉：free(E) / vision(F) / supports_search(G) / enabled(H) / is_default(J)
    dv_bool = DataValidation(
        type="list",
        formula1=bool_list,
        allow_blank=True,
        showDropDown=False,
    )
    dv_bool.error = bool_err
    dv_bool.errorTitle = bool_err_title
    ws.add_data_validation(dv_bool)
    for col in ("E", "F", "G", "H", "J"):
        dv_bool.add(f"{col}2:{col}1000")

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="models_template.xlsx"'},
    )


def _parse_bool(v) -> int:
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        return 1 if v else 0
    s = str(v or "").strip().lower()
    return 1 if s in ("1", "true", "yes", "是", "y", "on") else 0


def _has_chinese(s: str) -> bool:
    """是否含中文字符（CJK 统一表意文字基本区 + 扩展 A）。"""
    return any("一" <= c <= "鿿" or "㐀" <= c <= "䶿" for c in s)


@router.post("/models/import")
async def import_models(file: UploadFile = File(...), admin: dict = Depends(require_model_admin)):
    """从 xlsx（Excel 表格）批量导入模型：model_key 已存在则更新，否则新增。"""
    raw = await file.read()
    original_name = file.filename or "models_import.xlsx"
    username = admin.get("username") or "admin"
    if len(raw) > 10 * 1024 * 1024:
        # 超大文件不落盘，但同样写失败记录，同步到导入管理
        _write_transfer_meta(
            "import", username, original_name, len(raw), _XLSX_MIME,
            remark="模型数据导入：文件过大（上限 10MB）", status="failed",
        )
        raise HTTPException(status_code=400, detail="文件过大（上限 10MB）")

    # 先落盘源文件并写导入记录（无论后续解析是否成功，都保留导入痕迹）
    saved_path, record_id = _write_transfer_record(
        "import", username, original_name, raw, _XLSX_MIME,
        remark="模型数据导入", status="",
    )

    def _mark_failed(reason: str) -> None:
        execute(
            "UPDATE transfer_records SET status='failed', remark=? WHERE id=?",
            (f"模型数据导入：{reason}", record_id),
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    except Exception:
        _mark_failed("无法解析文件")
        raise HTTPException(status_code=400, detail="无法解析文件，请上传有效的 .xlsx 文件")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        _mark_failed("空文件")
        raise HTTPException(status_code=400, detail="空文件")

    # 表头兼容：去空格、小写
    header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    if "model_key" not in header:
        _mark_failed("缺少 model_key 列")
        raise HTTPException(status_code=400, detail="xlsx 缺少 model_key 列")
    if "name" not in header:
        _mark_failed("缺少 name 列")
        raise HTTPException(status_code=400, detail="xlsx 缺少 name 列")

    idx = {h: i for i, h in enumerate(header)}
    def cell(row, key: str, default=""):
        i = idx.get(key)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return "" if v is None else str(v).strip()

    created = 0
    updated = 0
    errors: list[str] = []
    row_errors: dict[int, str] = {}
    known_providers = set(_known_provider_ids())
    seen_names: set[str] = set()
    seen_names_en: set[str] = set()
    for n, row in enumerate(rows_iter, start=2):
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue
        model_key = cell(row, "model_key")
        name = cell(row, "name")
        if not model_key or not name:
            reason = "model_key/name 不能为空"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        name_en = cell(row, "name_en")
        if len(name_en) > 100:
            reason = "name_en 长度超限（100）"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        if name_en and _has_chinese(name_en):
            reason = "name_en 不能包含中文"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        provider = cell(row, "provider", "openai") or "openai"
        if provider not in known_providers:
            reason = "provider 不在数据字典"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        exists = fetch_one(
            "SELECT id FROM models WHERE provider = ? AND model_key = ?", (provider, model_key)
        )
        dup_name = fetch_one("SELECT id FROM models WHERE name = ?", (name,))
        if (dup_name and (not exists or dup_name["id"] != exists["id"])) or name in seen_names:
            reason = "模型名称已存在"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        dup_name_en = None
        if name_en:
            dup_name_en = fetch_one("SELECT id FROM models WHERE name_en = ? AND name_en <> ''", (name_en,))
        if (dup_name_en and (not exists or dup_name_en["id"] != exists["id"])) or name_en in seen_names_en:
            reason = "英文名称已存在"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        try:
            sort_order = int(float(cell(row, "sort_order", "0") or "0"))
        except ValueError:
            sort_order = 0
        free = _parse_bool(cell(row, "free", "0"))
        vision = _parse_bool(cell(row, "vision", "0"))
        supports_search = _parse_bool(cell(row, "supports_search", "1"))
        enabled = _parse_bool(cell(row, "enabled", "1"))
        is_default = _parse_bool(cell(row, "is_default", "0"))

        if exists:
            execute(
                "UPDATE models SET name=?, name_en=?, provider=?, free=?, vision=?, supports_search=?, enabled=?, sort_order=?, is_default=? WHERE model_key=? AND provider=?",
                (name, name_en, provider, free, vision, supports_search, enabled, sort_order, is_default, model_key, provider),
            )
            updated += 1
        else:
            execute(
                "INSERT INTO models (model_key, name, name_en, provider, free, vision, supports_search, enabled, sort_order, is_default, created_at)"
                " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (model_key, name, name_en, provider, free, vision, supports_search, enabled, sort_order, is_default, now_ms()),
            )
            created += 1
        seen_names.add(name)
        if name_en:
            seen_names_en.add(name_en)

    # 若导入导致多个默认模型，则收敛为唯一默认（优先 sort_order 最小且 enabled 的）
    if fetch_one("SELECT COUNT(*) AS n FROM models WHERE is_default = 1")["n"] > 1:
        execute("UPDATE models SET is_default = 0")
        d = fetch_one("SELECT id FROM models WHERE enabled = 1 ORDER BY sort_order, id LIMIT 1")
        if d:
            execute("UPDATE models SET is_default = 1 WHERE id = ?", (d["id"],))

    # 记录状态：全部失败 → failed；有行错误但有成功 → partial；否则 success
    if row_errors and (created + updated) == 0:
        status = "failed"
    elif row_errors:
        status = "partial"
    else:
        status = "success"

    # 失败/部分成功：下载产物替换为带「错误分析」列的标注文件（含全部源数据，是源文件的超集）
    if row_errors:
        try:
            ann_path, ann_size = _build_annotated_file(raw, row_errors, original_name)
            if os.path.isfile(saved_path):
                os.remove(saved_path)
            execute(
                "UPDATE transfer_records SET file_path=?, file_size=? WHERE id=?",
                (ann_path, ann_size, record_id),
            )
        except Exception:
            pass  # 标注失败保留源文件，不影响导入主流程

    execute("UPDATE transfer_records SET status=? WHERE id=?", (status, record_id))

    return {"ok": True, "created": created, "updated": updated, "errors": errors}


@router.get("/models/check")
def check_model_uniqueness(
    model_key: str = "",
    name: str = "",
    name_en: str = "",
    provider: str = "",
    exclude_id: int = 0,
    admin: dict = Depends(require_admin),
):
    """失焦轻量查重：model_key 同供应商内唯一；name / name_en 全局唯一。exclude_id 用于编辑时排除自身。"""
    key_exists = False
    name_exists = False
    if model_key and provider:
        r = fetch_one(
            "SELECT id FROM models WHERE provider = ? AND model_key = ?", (provider, model_key)
        )
        key_exists = bool(r and r["id"] != exclude_id)
    if name:
        r = fetch_one("SELECT id FROM models WHERE name = ?", (name,))
        name_exists = bool(r and r["id"] != exclude_id)
    name_en_exists = False
    if name_en:
        r = fetch_one("SELECT id FROM models WHERE name_en = ? AND name_en <> ''", (name_en,))
        name_en_exists = bool(r and r["id"] != exclude_id)
    return {"model_key_exists": key_exists, "name_exists": name_exists, "name_en_exists": name_en_exists}


@router.get("/models/{model_id}")
def get_model(model_id: int):
    row = fetch_one("SELECT * FROM models WHERE id = ?", (model_id,))
    if not row:
        raise HTTPException(status_code=404, detail="模型不存在")
    return dict(row)


@router.post("/models")
def create_model(body: ModelPayload, admin: dict = Depends(require_model_admin)):
    exists = fetch_one(
        "SELECT id FROM models WHERE provider = ? AND model_key = ?",
        (body.provider, body.model_key),
    )
    if exists:
        raise HTTPException(status_code=409, detail="同供应商下 model_key 已存在")
    if body.provider not in set(_known_provider_ids()):
        raise HTTPException(status_code=400, detail="provider 不在数据字典")
    dup_name = fetch_one("SELECT id FROM models WHERE name = ?", (body.name,))
    if dup_name:
        raise HTTPException(status_code=409, detail="模型名称已存在")
    if body.name_en:
        dup_name_en = fetch_one("SELECT id FROM models WHERE name_en = ? AND name_en <> ''", (body.name_en,))
        if dup_name_en:
            raise HTTPException(status_code=409, detail="英文名称已存在")
    if body.is_default and not body.enabled:
        raise HTTPException(status_code=400, detail="禁用模型不能设为默认模型")
    new_id = execute(
        "INSERT INTO models (model_key, name, name_en, provider, free, vision, supports_search, enabled, sort_order, is_default, created_at)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            body.model_key,
            body.name,
            body.name_en,
            body.provider,
            1 if body.free else 0,
            1 if body.vision else 0,
            1 if body.supports_search else 0,
            1 if body.enabled else 0,
            body.sort_order,
            1 if body.is_default else 0,
            now_ms(),
        ),
    )
    if body.is_default:
        # 设默认时清除其它默认，保证全局唯一
        transaction([
            ("UPDATE models SET is_default = 0 WHERE id <> ?", (new_id,)),
            ("UPDATE models SET is_default = 1 WHERE id = ?", (new_id,)),
        ])
    _log_operation(
        "model", new_id, admin,
        f"新增模型「{body.name}」（{body.model_key}）",
        f'Created model "{body.name}" ({body.model_key})',
    )
    return {"ok": True}


def _model_diff_logs(old: dict, body) -> list[tuple[str, str]]:
    """模型字段级变更日志：逐字段比较旧值→新值，返回 (中文, English) 列表。"""
    old = dict(old)  # 兼容 sqlite3.Row（fetch_one 返回）
    msgs: list[tuple[str, str]] = []

    def yn(v) -> tuple[str, str]:
        return ("是" if v else "否", "Yes" if v else "No")

    def en_txt(v) -> tuple[str, str]:
        return ("启用" if v else "禁用", "enabled" if v else "disabled")

    if old["model_key"] != body.model_key:
        msgs.append((f'模型标识从「{old["model_key"]}」改成「{body.model_key}」',
                     f'Model key changed from "{old["model_key"]}" to "{body.model_key}"'))
    if old["name"] != body.name:
        msgs.append((f'名称从「{old["name"]}」改成「{body.name}」',
                     f'Name changed from "{old["name"]}" to "{body.name}"'))
    old_name_en = old.get("name_en") or ""
    new_name_en = body.name_en or ""
    if old_name_en != new_name_en:
        msgs.append((f'英文名称从「{old_name_en}」改成「{new_name_en}」',
                     f'English name changed from "{old_name_en}" to "{new_name_en}"'))
    if old["provider"] != body.provider:
        msgs.append((f'提供商从「{old["provider"]}」改成「{body.provider}」',
                     f'Provider changed from "{old["provider"]}" to "{body.provider}"'))
    if bool(old["free"]) != bool(body.free):
        o, oe = yn(bool(old["free"])); n, ne = yn(body.free)
        msgs.append((f'免费从「{o}」改成「{n}」', f'Free changed from "{oe}" to "{ne}"'))
    if bool(old["vision"]) != bool(body.vision):
        o, oe = yn(bool(old["vision"])); n, ne = yn(body.vision)
        msgs.append((f'视觉从「{o}」改成「{n}」', f'Vision changed from "{oe}" to "{ne}"'))
    if bool(old["supports_search"]) != bool(body.supports_search):
        o, oe = yn(bool(old["supports_search"])); n, ne = yn(body.supports_search)
        msgs.append((f'联网从「{o}」改成「{n}」', f'Web search changed from "{oe}" to "{ne}"'))
    if bool(old["enabled"]) != bool(body.enabled):
        o, oe = en_txt(bool(old["enabled"])); n, ne = en_txt(body.enabled)
        msgs.append((f'状态从「{o}」改成「{n}」', f'Status changed from "{oe}" to "{ne}"'))
    if int(old["sort_order"]) != int(body.sort_order):
        msgs.append((f'排序从「{old["sort_order"]}」改成「{body.sort_order}」',
                     f'Sort order changed from "{old["sort_order"]}" to "{body.sort_order}"'))
    if bool(old["is_default"]) != bool(body.is_default):
        o, oe = yn(bool(old["is_default"])); n, ne = yn(body.is_default)
        msgs.append((f'默认从「{o}」改成「{n}」', f'Default changed from "{oe}" to "{ne}"'))
    return msgs


@router.put("/models/{model_id}")
def update_model(model_id: int, body: ModelPayload, admin: dict = Depends(require_model_admin)):
    row = fetch_one("SELECT * FROM models WHERE id = ?", (model_id,))
    if not row:
        raise HTTPException(status_code=404, detail="模型不存在")

    if body.provider not in set(_known_provider_ids()):
        raise HTTPException(status_code=400, detail="provider 不在数据字典")
    key_dup = fetch_one(
        "SELECT id FROM models WHERE provider = ? AND model_key = ?",
        (body.provider, body.model_key),
    )
    if key_dup and key_dup["id"] != model_id:
        raise HTTPException(status_code=409, detail="同供应商下 model_key 已存在")
    dup_name = fetch_one("SELECT id FROM models WHERE name = ?", (body.name,))
    if dup_name and dup_name["id"] != model_id:
        raise HTTPException(status_code=409, detail="模型名称已存在")
    if body.name_en:
        dup_name_en = fetch_one("SELECT id FROM models WHERE name_en = ? AND name_en <> ''", (body.name_en,))
        if dup_name_en and dup_name_en["id"] != model_id:
            raise HTTPException(status_code=409, detail="英文名称已存在")

    effective_enabled = body.enabled if body.enabled is not None else bool(row["enabled"])

    # 字段级变更日志（旧值 → 新值）
    log_msgs = _model_diff_logs(row, body)

    if not body.is_default:
        # 未请求设为默认：保护与默认相关的约束
        if body.enabled is False and row["is_default"]:
            raise HTTPException(status_code=400, detail="默认模型不能被禁用，请先指定其它默认模型")
        if body.is_default is False and row["is_default"]:
            raise HTTPException(status_code=400, detail="必须保留一个默认模型，请先指定其它默认模型")
        # 普通更新，不动 is_default
        execute(
            "UPDATE models SET model_key=?, name=?, name_en=?, provider=?, free=?, vision=?, supports_search=?, enabled=?, sort_order=? WHERE id=?",
            (
                body.model_key,
                body.name,
                body.name_en,
                body.provider,
                1 if body.free else 0,
                1 if body.vision else 0,
                1 if body.supports_search else 0,
                1 if body.enabled else 0,
                body.sort_order,
                model_id,
            ),
        )
        for zh, en in log_msgs:
            _log_operation("model", model_id, admin, zh, en)
        return {"ok": True}

    # 请求设为默认：禁用模型不允许
    if not effective_enabled:
        raise HTTPException(status_code=400, detail="禁用模型不能设为默认模型")
    # 清除其它默认 + 更新本行字段并置为默认，保证全局唯一
    params = (
        body.model_key,
        body.name,
        body.name_en,
        body.provider,
        1 if body.free else 0,
        1 if body.vision else 0,
        1 if body.supports_search else 0,
        1 if body.enabled else 0,
        body.sort_order,
        model_id,
    )
    transaction([
        ("UPDATE models SET is_default = 0 WHERE id <> ?", (model_id,)),
        (
            "UPDATE models SET model_key=?, name=?, name_en=?, provider=?, free=?, vision=?, supports_search=?, enabled=?, sort_order=?, is_default=1 WHERE id=?",
            params,
        ),
    ])
    for zh, en in log_msgs:
        _log_operation("model", model_id, admin, zh, en)
    return {"ok": True}


@router.delete("/models/{model_id}")
def delete_model(model_id: int, admin: dict = Depends(require_model_admin)):
    row = fetch_one("SELECT id, name, model_key FROM models WHERE id = ?", (model_id,))
    if not row:
        raise HTTPException(status_code=404, detail="模型不存在")
    execute("DELETE FROM models WHERE id = ?", (model_id,))
    _log_operation(
        "model", model_id, admin,
        f"删除模型「{row['name']}」（{row['model_key']}）",
        f'Deleted model "{row["name"]}" ({row["model_key"]})',
    )
    return {"ok": True}


@router.get("/settings")
def list_settings(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    search: str = "",
    enabled: str = "",
    sort: str = "",
    order: str = "asc",
):
    offset = (page - 1) * page_size
    clauses: list[str] = []
    params: list = []
    if search:
        clauses.append("(key LIKE ? OR value LIKE ? OR remark LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
    if enabled:
        clauses.append("enabled = ?")
        params.append(1 if enabled.strip().lower() == "true" else 0)
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    sort_columns = {"key": "key", "value": "value", "remark": "remark", "enabled": "enabled"}
    order_by = "key"
    if sort in sort_columns:
        direction = "ASC" if order.strip().lower() == "asc" else "DESC"
        order_by = f"{sort_columns[sort]} {direction}"
    total = fetch_one(f"SELECT COUNT(*) AS n FROM settings {where}", params)["n"]
    rows = fetch_all(
        f"SELECT * FROM settings {where} ORDER BY {order_by} LIMIT ? OFFSET ?",
        (*params, page_size, offset),
    )
    return {"items": [dict(r) for r in rows], "total": total, "page": page, "pageSize": page_size}


@router.patch("/settings/{key}")
def update_setting(key: str, body: SettingsPayload, admin: dict = Depends(require_settings_admin)):
    key = _validate_setting_key(key)
    row = fetch_one("SELECT * FROM settings WHERE key = ?", (key,))
    if row is None:
        # 不存在视为新增（INSERT OR IGNORE 兼容已存在的情况）
        execute(
            "INSERT OR IGNORE INTO settings (key, value, remark, enabled) VALUES (?, '', '', 1)",
            (key,),
        )
        old = {"value": "", "remark": "", "enabled": 1}
        _log_setting(key, admin, f"新增配置项「{key}」", f'Created setting "{key}"')
    else:
        old = dict(row)
    updates: list[str] = []
    params: list = []
    log_msgs: list[tuple[str, str]] = []  # (中文, English)
    if body.value is not None:
        updates.append("value = ?")
        params.append(body.value)
        if old["value"] != body.value:
            log_msgs.append(
                (
                    f'配置值从「{old["value"]}」改成「{body.value}」',
                    f'Value changed from "{old["value"]}" to "{body.value}"',
                )
            )
    if body.remark is not None:
        updates.append("remark = ?")
        params.append(body.remark)
        if old["remark"] != body.remark:
            log_msgs.append(
                (
                    f'备注从「{old["remark"]}」改成「{body.remark}」',
                    f'Remark changed from "{old["remark"]}" to "{body.remark}"',
                )
            )
    if body.enabled is not None:
        updates.append("enabled = ?")
        params.append(1 if body.enabled else 0)
        old_enabled = bool(old["enabled"])
        if old_enabled != body.enabled:
            old_txt = "开启" if old_enabled else "禁用"
            new_txt = "开启" if body.enabled else "禁用"
            old_en = "enabled" if old_enabled else "disabled"
            new_en = "enabled" if body.enabled else "disabled"
            log_msgs.append(
                (f"状态从「{old_txt}」改成「{new_txt}」", f'Status changed from "{old_en}" to "{new_en}"')
            )
    if updates:
        params.append(key)
        execute(f"UPDATE settings SET {', '.join(updates)} WHERE key = ?", params)
    for zh, en in log_msgs:
        _log_setting(key, admin, zh, en)
    return {"ok": True}


@router.delete("/settings/{key}")
def delete_setting(key: str, admin: dict = Depends(require_settings_admin)):
    key = _validate_setting_key(key)
    row = fetch_one("SELECT * FROM settings WHERE key = ?", (key,))
    if row is None:
        raise HTTPException(status_code=404, detail="配置项不存在")
    if row["enabled"]:
        raise HTTPException(status_code=400, detail="启用状态的配置项不能删除，请先禁用")
    execute("DELETE FROM settings WHERE key = ?", (key,))
    _log_setting(key, admin, f"删除配置项「{key}」", f'Deleted setting "{key}"')
    return {"ok": True}


# 配置项可导入导出的字段顺序（与 xlsx 表头一致）
_SETTING_FIELDS = ["key", "value", "remark", "enabled"]


def _setting_rows() -> list[list]:
    rows = fetch_all("SELECT * FROM settings ORDER BY key")
    return [
        [r["key"], "" if r["value"] is None else r["value"], "" if r["remark"] is None else r["remark"], 1 if r["enabled"] else 0]
        for r in rows
    ]


@router.get("/settings/export")
def export_settings(admin: dict = Depends(require_settings_admin)):
    """导出全部配置项为 xlsx。导出不存产物：仅写行为记录。"""
    username = admin.get("username") or "admin"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "settings"
    ws.append(_SETTING_FIELDS)
    for row in _setting_rows():
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    filename = f"settings_{now_ms()}.xlsx"
    _write_transfer_meta(
        "export", username, filename, len(data), _XLSX_MIME,
        remark="配置数据导出", status="success",
    )
    return StreamingResponse(
        iter([data]),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/settings/template")
def download_settings_template(request: Request, admin: dict = Depends(require_settings_admin)):
    """下载配置项导入模板：英文字段表头 + 本地化示例行 + enabled 布尔下拉。"""
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "settings"
    ws.append(_SETTING_FIELDS)
    ws.append(["example_key", "example value", "example remark", "yes"] if is_en
               else ["example_key", "示例值", "示例说明", "是"])
    dv = DataValidation(
        type="list",
        formula1='"yes,no"' if is_en else '"是,否"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = 'Please choose "yes" or "no"' if is_en else "请选择「是」或「否」"
    dv.errorTitle = "Invalid value" if is_en else "非法值"
    ws.add_data_validation(dv)
    dv.add("D2:D1000")
    return _dim_xlsx_response(wb, "settings_template.xlsx")


@router.post("/settings/import")
async def import_settings(file: UploadFile = File(...), admin: dict = Depends(require_settings_admin)):
    """从 xlsx 批量导入配置项：key 已存在则更新，否则新增。"""
    raw = await file.read()
    original_name = file.filename or "settings_import.xlsx"
    username = admin.get("username") or "admin"
    if len(raw) > 10 * 1024 * 1024:
        _write_transfer_meta(
            "import", username, original_name, len(raw), _XLSX_MIME,
            remark="配置数据导入：文件过大（上限 10MB）", status="failed",
        )
        raise HTTPException(status_code=400, detail="文件过大（上限 10MB）")

    saved_path, record_id = _write_transfer_record(
        "import", username, original_name, raw, _XLSX_MIME,
        remark="配置数据导入", status="",
    )

    def _mark_failed(reason: str) -> None:
        execute(
            "UPDATE transfer_records SET status='failed', remark=? WHERE id=?",
            (f"配置数据导入：{reason}", record_id),
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    except Exception:
        _mark_failed("无法解析文件")
        raise HTTPException(status_code=400, detail="无法解析文件，请上传有效的 .xlsx 文件")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        _mark_failed("空文件")
        raise HTTPException(status_code=400, detail="空文件")

    header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    if "key" not in header:
        _mark_failed("缺少 key 列")
        raise HTTPException(status_code=400, detail="xlsx 缺少 key 列")

    idx = {h: i for i, h in enumerate(header)}

    def cell(row, key: str, default=""):
        i = idx.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return default
        return row[i]

    created = 0
    updated = 0
    errors: list[str] = []
    row_errors: dict[int, str] = {}
    seen_keys: set[str] = set()

    for n, row in enumerate(rows_iter, start=2):
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue
        key = str(cell(row, "key", "")).strip()
        if not key:
            reason = "key 不能为空"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        if not (1 <= len(key) < 64) or not _SETTING_KEY_RE.match(key):
            reason = "key 格式非法（仅英文字母/下划线，长度 1-63）"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        if key in seen_keys:
            reason = f"文件内键名重复：{key}"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        value = str(cell(row, "value", "") or "")
        remark = str(cell(row, "remark", "") or "")
        if len(value) >= 5000:
            reason = "value 长度超限（5000）"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        if len(remark) > 255:
            reason = "remark 长度超限（255）"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        enabled = _parse_bool(cell(row, "enabled", "1"))

        exists = fetch_one("SELECT key FROM settings WHERE key = ?", (key,))
        if exists:
            execute(
                "UPDATE settings SET value = ?, remark = ?, enabled = ? WHERE key = ?",
                (value, remark, enabled, key),
            )
            _log_setting(key, admin, f"导入更新配置项「{key}」", f'Imported update to setting "{key}"')
            updated += 1
        else:
            execute(
                "INSERT INTO settings (key, value, remark, enabled) VALUES (?, ?, ?, ?)",
                (key, value, remark, enabled),
            )
            _log_setting(key, admin, f"导入新增配置项「{key}」", f'Imported new setting "{key}"')
            created += 1
        seen_keys.add(key)

    if row_errors and (created + updated) == 0:
        status = "failed"
    elif row_errors:
        status = "partial"
    else:
        status = "success"

    if row_errors:
        try:
            ann_path, ann_size = _build_annotated_file(raw, row_errors, original_name)
            if os.path.isfile(saved_path):
                os.remove(saved_path)
            execute(
                "UPDATE transfer_records SET file_path=?, file_size=? WHERE id=?",
                (ann_path, ann_size, record_id),
            )
        except Exception:
            pass

    execute("UPDATE transfer_records SET status=? WHERE id=?", (status, record_id))
    return {"ok": True, "created": created, "updated": updated, "errors": errors}


@router.get("/settings/{key}/logs", response_model=SettingLogList)
def list_setting_logs(
    key: str,
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    admin: dict = Depends(require_settings_admin),
):
    offset = (page - 1) * page_size
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    total = fetch_one(
        "SELECT COUNT(*) AS n FROM setting_logs WHERE setting_key = ?", (key,)
    )["n"]
    rows = fetch_all(
        "SELECT * FROM setting_logs WHERE setting_key = ? ORDER BY created_at DESC, id DESC LIMIT ? OFFSET ?",
        (key, page_size, offset),
    )
    items = []
    for r in rows:
        d = dict(r)
        if is_en:
            d["content"] = d.get("content_en") or d["content"]
        items.append(d)
    return {"items": items, "total": total, "page": page, "pageSize": page_size}


@router.get("/operation-logs", response_model=OperationLogList)
def list_operation_logs(
    entity: str,
    entity_id: int,
    request: Request,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
):
    """通用操作日志：按实体类型 + 实体 id 分页查询（模型/维表取值/用户共用）。
    content 按 Accept-Language 返回中文或英文。"""
    offset = (page - 1) * page_size
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    total = fetch_one(
        "SELECT COUNT(*) AS n FROM admin_operation_logs WHERE entity = ? AND entity_id = ?",
        (entity, entity_id),
    )["n"]
    rows = fetch_all(
        "SELECT * FROM admin_operation_logs WHERE entity = ? AND entity_id = ? "
        "ORDER BY created_at DESC, id DESC LIMIT ? OFFSET ?",
        (entity, entity_id, page_size, offset),
    )
    items = []
    for r in rows:
        d = dict(r)
        if is_en:
            d["content"] = d.get("content_en") or d["content"]
        items.append(d)
    return {"items": items, "total": total, "page": page, "pageSize": page_size}


# ============ 系统设置辅助（key 校验 / 操作日志） ============

_SETTING_KEY_RE = re.compile(r"^[A-Za-z_]+$")


def _validate_setting_key(key: str) -> str:
    key = key.strip()
    if not (1 <= len(key) < 64):
        raise HTTPException(status_code=422, detail="键名长度需在 1-63 之间")
    if not _SETTING_KEY_RE.match(key):
        raise HTTPException(status_code=422, detail="键名只能包含英文字母和下划线")
    return key


def _log_setting(key: str, admin: dict, content: str, content_en: str = "") -> None:
    execute(
        "INSERT INTO setting_logs (setting_key, content, content_en, operator, created_at) VALUES (?, ?, ?, ?, ?)",
        (key, content, content_en, admin.get("username") or "", now_ms()),
    )


def _log_operation(entity: str, entity_id: int, admin: dict, content: str, content_en: str = "") -> None:
    """写入通用操作日志（模型 / 维表取值 / 用户共用）。"""
    execute(
        "INSERT INTO admin_operation_logs (entity, entity_id, content, content_en, operator, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        (entity, entity_id, content, content_en, admin.get("username") or "", now_ms()),
    )


# ============ 通用维表（dim_tables / dim_values） ============


def _dim_table_out(row) -> DimTableOut:
    count = fetch_one(
        "SELECT COUNT(*) AS n FROM dim_values WHERE table_id = ?", (row["id"],)
    )["n"]
    return DimTableOut(
        id=row["id"],
        code=row["code"],
        name=row["name"],
        description=row["description"] or "",
        sort_order=row["sort_order"] or 0,
        value_count=count,
        created_at=row["created_at"],
        updated_at=row["updated_at"],
        updated_by=row["updated_by"] or "",
    )


@router.get("/dim-tables")
def list_dim_tables(admin: dict = Depends(require_settings_admin)):
    rows = fetch_all("SELECT * FROM dim_tables ORDER BY sort_order, id")
    return [_dim_table_out(r) for r in rows]


@router.post("/dim-tables")
def create_dim_table(body: DimTableCreate, admin: dict = Depends(require_settings_admin)):
    existing = fetch_one("SELECT id FROM dim_tables WHERE code = ?", (body.code,))
    if existing:
        raise HTTPException(status_code=400, detail="维表编码已存在")
    ts = now_ms()
    cur = execute(
        "INSERT INTO dim_tables (code, name, description, sort_order, created_at, updated_at, updated_by) "
        "VALUES (?, ?, ?, 0, ?, ?, ?)",
        (body.code, body.name, body.description, ts, ts, admin.get("username") or "admin"),
    )
    row = fetch_one("SELECT * FROM dim_tables WHERE id = ?", (cur,))
    return _dim_table_out(row)


@router.put("/dim-tables/{table_id}")
def update_dim_table(table_id: int, body: DimTableUpdate, admin: dict = Depends(require_settings_admin)):
    row = fetch_one("SELECT id FROM dim_tables WHERE id = ?", (table_id,))
    if not row:
        raise HTTPException(status_code=404, detail="维表不存在")
    updates: list[str] = []
    params: list = []
    if body.name is not None:
        updates.append("name = ?")
        params.append(body.name)
    if body.description is not None:
        updates.append("description = ?")
        params.append(body.description)
    if body.sort_order is not None:
        updates.append("sort_order = ?")
        params.append(body.sort_order)
    if updates:
        updates.append("updated_at = ?")
        params.append(now_ms())
        updates.append("updated_by = ?")
        params.append(admin.get("username") or "admin")
        params.append(table_id)
        execute(f"UPDATE dim_tables SET {', '.join(updates)} WHERE id = ?", params)
    return {"ok": True}


@router.delete("/dim-tables/{table_id}")
def delete_dim_table(table_id: int, admin: dict = Depends(require_settings_admin)):
    row = fetch_one("SELECT id FROM dim_tables WHERE id = ?", (table_id,))
    if not row:
        raise HTTPException(status_code=404, detail="维表不存在")
    execute("DELETE FROM dim_tables WHERE id = ?", (table_id,))
    return {"ok": True}


@router.get("/dim-tables/{table_id}/values")
def list_dim_values(
    table_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    search: str = "",
    enabled: str = "",
    sort: str = "",
    order: str = "asc",
    admin: dict = Depends(require_settings_admin),
):
    table = fetch_one("SELECT id FROM dim_tables WHERE id = ?", (table_id,))
    if not table:
        raise HTTPException(status_code=404, detail="维表不存在")
    offset = (page - 1) * page_size
    clauses = ["table_id = ?"]
    params: list = [table_id]
    if search:
        clauses.append("(code LIKE ? OR name LIKE ? OR name_en LIKE ? OR remark LIKE ?)")
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])
    if enabled:
        clauses.append("enabled = ?")
        params.append(1 if enabled.strip().lower() == "true" else 0)
    where = "WHERE " + " AND ".join(clauses)
    sort_columns = {
        "code": "code",
        "name": "name",
        "name_en": "name_en",
        "sort_order": "sort_order",
        "enabled": "enabled",
        "remark": "remark",
    }
    order_by = "sort_order, id"
    if sort in sort_columns:
        direction = "ASC" if order.strip().lower() == "asc" else "DESC"
        order_by = f"{sort_columns[sort]} {direction}, id"
    total = fetch_one(f"SELECT COUNT(*) AS n FROM dim_values {where}", params)["n"]
    rows = fetch_all(
        f"SELECT * FROM dim_values {where} ORDER BY {order_by} LIMIT ? OFFSET ?",
        (*params, page_size, offset),
    )
    # 非 super/system 管理员：api_key 脱敏回显（中间星号），避免明文扩散
    privileged = admin.get("role") in (ROLE_SUPER_ADMIN, ROLE_SYSTEM_ADMIN)
    items = []
    for r in rows:
        d = dict(r)
        if not privileged:
            d["api_key"] = _mask_secret(d.get("api_key"))
        items.append(DimValueOut(**d))
    return DimValueList(items=items, total=total, page=page, pageSize=page_size)


# 英文名称禁止出现汉字（CJK 统一表意文字）
_CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]")


def _is_model_provider_table(table_id: int) -> bool:
    t = fetch_one("SELECT code FROM dim_tables WHERE id = ?", (table_id,))
    return bool(t and t["code"] == "model_provider")


def _validate_dim_name_en(table_id: int, name_en: str | None, *, required: bool) -> str | None:
    """模型供应商维表的英文名校验：必填 + 不允许中文。返回错误文案或 None。"""
    if not _is_model_provider_table(table_id):
        return None
    if name_en is None:
        return None
    val = name_en.strip()
    if required and not val:
        return "模型供应商维表的英文名称不能为空"
    if _CJK_RE.search(val):
        return "英文名称不允许包含中文"
    return None


# 连续 4 个及以上 * 视为「掩码回显」：非特权角色看到掩码后若原样回传，不得覆盖真实密钥
_MASK_RE = re.compile(r"\*{4,}")


def _is_masked_secret(s: str | None) -> bool:
    return bool(s) and bool(_MASK_RE.search(s or ""))


def _mask_secret(s: str | None) -> str:
    """非特权角色列表回显脱敏：保留前 3 后 4，中间固定 8 个 *；过短则全掩码。"""
    if not s:
        return ""
    if len(s) <= 8:
        return "********"
    return f"{s[:3]}********{s[-4:]}"


def _provider_needs_key(code: str | None) -> bool:
    """ollama 为本地部署、无需 API 密钥，启用规则予以豁免。"""
    return (code or "") != "ollama"


def _validate_provider_enable(
    table_id: int, code: str | None, api_key: str | None, enabled: bool
) -> str | None:
    """模型供应商维表：要启用（enabled=True）必须已配置真实 api_key（ollama 豁免）。返回错误文案或 None。"""
    if not _is_model_provider_table(table_id):
        return None
    if not enabled or not _provider_needs_key(code):
        return None
    if not (api_key or "").strip() or _is_masked_secret(api_key):
        return "API 密钥是启用提供商的必要条件，请先填写 API 密钥"
    return None


@router.post("/dim-tables/{table_id}/values")
def create_dim_value(
    table_id: int, body: DimValueCreate, admin: dict = Depends(require_settings_admin)
):
    table = fetch_one("SELECT id FROM dim_tables WHERE id = ?", (table_id,))
    if not table:
        raise HTTPException(status_code=404, detail="维表不存在")
    err = _validate_dim_name_en(table_id, body.name_en, required=True)
    if err:
        raise HTTPException(status_code=422, detail=err)
    # 掩码回显原样提交视为未填写，避免覆盖真实密钥
    api_key = "" if _is_masked_secret(body.api_key) else (body.api_key or "").strip()
    enable_err = _validate_provider_enable(table_id, body.code, api_key, body.enabled)
    if enable_err:
        raise HTTPException(status_code=422, detail=enable_err)
    dup = fetch_one(
        "SELECT id FROM dim_values WHERE table_id = ? AND code = ?", (table_id, body.code)
    )
    if dup:
        raise HTTPException(status_code=400, detail="该编码在当前维表已存在")
    ts = now_ms()
    value_id = execute(
        "INSERT INTO dim_values (table_id, code, name, name_en, api_key, sort_order, enabled, remark, created_at, updated_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (
            table_id,
            body.code,
            body.name,
            (body.name_en or "").strip(),
            api_key,
            body.sort_order,
            1 if body.enabled else 0,
            body.remark,
            ts,
            ts,
        ),
    )
    _log_operation(
        "dim_value", value_id, admin,
        f"新增取值「{body.name}」（{body.code}）",
        f'Created value "{body.name}" ({body.code})',
    )
    return {"ok": True}


def _dim_value_diff_logs(old: dict, body) -> list[tuple[str, str]]:
    """维表取值字段级变更日志：逐字段比较旧值→新值，返回 (中文, English) 列表。"""
    old = dict(old)  # 兼容 sqlite3.Row（fetch_one 返回）
    msgs: list[tuple[str, str]] = []
    if body.code is not None and old["code"] != body.code:
        msgs.append((f'编码从「{old["code"]}」改成「{body.code}」',
                     f'Code changed from "{old["code"]}" to "{body.code}"'))
    if body.name is not None and old["name"] != body.name:
        msgs.append((f'名称从「{old["name"]}」改成「{body.name}」',
                     f'Name changed from "{old["name"]}" to "{body.name}"'))
    if body.name_en is not None:
        new_ne = body.name_en.strip()
        if (old.get("name_en") or "") != new_ne:
            msgs.append((f'英文名称从「{old.get("name_en") or ""}」改成「{new_ne}」',
                         f'English name changed from "{old.get("name_en") or ""}" to "{new_ne}"'))
    if body.api_key is not None and not _is_masked_secret(body.api_key):
        new_key = body.api_key.strip()
        if (old.get("api_key") or "") != new_key:
            # 密钥敏感，日志不记录具体值
            msgs.append(("API密钥已更新", "API key updated"))
    if body.sort_order is not None and int(old.get("sort_order") or 0) != int(body.sort_order):
        msgs.append((f'排序从「{old.get("sort_order")}」改成「{body.sort_order}」',
                     f'Sort order changed from "{old.get("sort_order")}" to "{body.sort_order}"'))
    if body.enabled is not None and bool(old["enabled"]) != bool(body.enabled):
        o = "启用" if bool(old["enabled"]) else "禁用"
        n = "启用" if body.enabled else "禁用"
        oe = "enabled" if bool(old["enabled"]) else "disabled"
        ne = "enabled" if body.enabled else "disabled"
        msgs.append((f'状态从「{o}」改成「{n}」', f'Status changed from "{oe}" to "{ne}"'))
    if body.remark is not None and (old.get("remark") or "") != body.remark:
        msgs.append((f'备注从「{old.get("remark") or ""}」改成「{body.remark}」',
                     f'Remark changed from "{old.get("remark") or ""}" to "{body.remark}"'))
    return msgs


@router.put("/dim-tables/{table_id}/values/{value_id}")
def update_dim_value(
    table_id: int,
    value_id: int,
    body: DimValueUpdate,
    admin: dict = Depends(require_settings_admin),
):
    row = fetch_one(
        "SELECT * FROM dim_values WHERE id = ? AND table_id = ?",
        (value_id, table_id),
    )
    if not row:
        raise HTTPException(status_code=404, detail="维表取值不存在")
    updates: list[str] = []
    params: list = []
    if body.code is not None:
        dup = fetch_one(
            "SELECT id FROM dim_values WHERE table_id = ? AND code = ? AND id != ?",
            (table_id, body.code, value_id),
        )
        if dup:
            raise HTTPException(status_code=400, detail="该编码在当前维表已存在")
        updates.append("code = ?")
        params.append(body.code)
    if body.name is not None:
        updates.append("name = ?")
        params.append(body.name)
    if body.name_en is not None:
        err = _validate_dim_name_en(table_id, body.name_en, required=True)
        if err:
            raise HTTPException(status_code=422, detail=err)
        updates.append("name_en = ?")
        params.append(body.name_en.strip())
    # api_key：掩码回显原样提交时跳过，绝不用星号覆盖真实密钥
    if body.api_key is not None and not _is_masked_secret(body.api_key):
        updates.append("api_key = ?")
        params.append(body.api_key.strip())
    if body.sort_order is not None:
        updates.append("sort_order = ?")
        params.append(body.sort_order)
    if body.enabled is not None:
        updates.append("enabled = ?")
        params.append(1 if body.enabled else 0)
    if body.remark is not None:
        updates.append("remark = ?")
        params.append(body.remark)
    if updates:
        # 启用校验：合并「传入值 + 现值」得到最终状态，要启用就必须有真实密钥（ollama 豁免）
        final_code = body.code if body.code is not None else row["code"]
        if body.api_key is not None and not _is_masked_secret(body.api_key):
            final_key = body.api_key.strip()
        else:
            final_key = row["api_key"]
        final_enabled = body.enabled if body.enabled is not None else bool(row["enabled"])
        enable_err = _validate_provider_enable(table_id, final_code, final_key, final_enabled)
        if enable_err:
            raise HTTPException(status_code=422, detail=enable_err)
        updates.append("updated_at = ?")
        params.append(now_ms())
        params.append(value_id)
        execute(f"UPDATE dim_values SET {', '.join(updates)} WHERE id = ?", params)
        # 字段级变更日志（旧值 → 新值）
        for zh, en in _dim_value_diff_logs(row, body):
            _log_operation("dim_value", value_id, admin, zh, en)
    return {"ok": True}


@router.delete("/dim-tables/{table_id}/values/{value_id}")
def delete_dim_value(
    table_id: int, value_id: int, admin: dict = Depends(require_settings_admin)
):
    row = fetch_one(
        "SELECT id, name FROM dim_values WHERE id = ? AND table_id = ?", (value_id, table_id)
    )
    if not row:
        raise HTTPException(status_code=404, detail="维表取值不存在")
    execute("DELETE FROM dim_values WHERE id = ?", (value_id,))
    _log_operation(
        "dim_value", value_id, admin,
        f"删除取值「{row['name']}」",
        f'Deleted value "{row["name"]}"',
    )
    return {"ok": True}


@router.get("/dim-tables/by-code/{code}/values")
def list_dim_values_by_code(
    code: str, request: Request, admin: dict = Depends(require_settings_admin)
):
    """下拉专用：返回某维表启用取值的 [{code, name}]，name 按 Accept-Language 返回中文或英文。"""
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    table = fetch_one("SELECT id FROM dim_tables WHERE code = ?", (code,))
    if not table:
        return []
    rows = fetch_all(
        "SELECT code, name, name_en FROM dim_values WHERE table_id = ? AND enabled = 1 ORDER BY sort_order, id",
        (table["id"],),
    )
    if is_en:
        return [{"code": r["code"], "name": (r["name_en"] or r["name"])} for r in rows]
    return [{"code": r["code"], "name": r["name"]} for r in rows]


# ============ 维表取值：公共导入 / 导出 / 模板（字段规格驱动，新增维表自动复用） ============

def _dim_table_or_404(table_id: int) -> dict:
    t = fetch_one("SELECT * FROM dim_tables WHERE id = ?", (table_id,))
    if not t:
        raise HTTPException(status_code=404, detail="维表不存在")
    return dict(t)


def _dim_value_field_spec(table_code: str) -> list[dict]:
    """维表取值的列规格，统一驱动「导出表头 / 模板 / 导入解析与校验 / upsert」。
    以后新增维表：通用列自动覆盖；若有专属列，只需在此按 table_code 扩展，接口无需重写。
    字段含义：required 必填；max_len 长度上限；no_cjk 禁中文；is_int 整型；is_bool 布尔。"""
    fields: list[dict] = [
        {"key": "code", "required": True, "max_len": 64},
        {"key": "name", "required": True, "max_len": 128},
    ]
    if table_code == "model_provider":
        fields.append({"key": "name_en", "required": True, "max_len": 128, "no_cjk": True})
        fields.append({"key": "api_key", "max_len": 512})
    fields.extend(
        [
            {"key": "sort_order", "is_int": True, "default": 0},
            {"key": "enabled", "is_bool": True, "default": 1},
            {"key": "remark", "max_len": 255},
        ]
    )
    return fields


def _dim_xlsx_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    return StreamingResponse(
        iter([data]),
        media_type=_XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/dim-tables/{table_id}/export")
def export_dim_values(
    table_id: int, request: Request, admin: dict = Depends(require_settings_admin)
):
    """导出某维表全部取值为 xlsx（列由字段规格决定，model_provider 含 name_en/api_key）。"""
    table = _dim_table_or_404(table_id)
    spec = _dim_value_field_spec(table["code"])
    keys = [f["key"] for f in spec]
    rows = fetch_all(
        "SELECT * FROM dim_values WHERE table_id = ? ORDER BY sort_order, id", (table_id,)
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "values"
    ws.append(keys)
    for r in rows:
        ws.append(["" if r[k] is None else r[k] for k in keys])
    username = admin.get("username") or "admin"
    fname = f"dim_{table['code']}_{now_ms()}.xlsx"
    resp = _dim_xlsx_response(wb, fname)
    # 导出行为记录（不落盘产物）
    buf_size = sum(1 for _ in rows)
    _write_transfer_meta(
        "export", username, fname, buf_size, _XLSX_MIME,
        remark=f"维表「{table['name']}」导出",
    )
    return resp


@router.get("/dim-tables/{table_id}/template")
def download_dim_template(
    table_id: int, request: Request, admin: dict = Depends(require_settings_admin)
):
    """下载维表取值导入模板：英文字段表头 + 本地化示例行 + enabled 布尔下拉。"""
    table = _dim_table_or_404(table_id)
    is_en = "en" in (request.headers.get("accept-language") or "").lower()
    spec = _dim_value_field_spec(table["code"])
    keys = [f["key"] for f in spec]
    is_provider = table["code"] == "model_provider"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "values"
    ws.append(keys)
    if is_provider:
        example = (
            ["example_code", "Example Name", "Example EN Name", "sk-xxxx", 0, "yes", ""]
            if is_en
            else ["example_code", "示例名称", "Example Name", "sk-xxxx", 0, "是", ""]
        )
    else:
        example = (
            ["example_code", "Example Name", 0, "yes", ""]
            if is_en
            else ["example_code", "示例名称", 0, "是", ""]
        )
    ws.append(example)

    # enabled 列布尔下拉
    enabled_idx = keys.index("enabled") + 1
    letter = get_column_letter(enabled_idx)
    dv = DataValidation(
        type="list",
        formula1='"yes,no"' if is_en else '"是,否"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = 'Please choose "yes" or "no"' if is_en else "请选择「是」或「否」"
    dv.errorTitle = "Invalid value" if is_en else "非法值"
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}1000")

    return _dim_xlsx_response(wb, f"dim_{table['code']}_template.xlsx")


@router.post("/dim-tables/{table_id}/import")
async def import_dim_values(
    table_id: int, file: UploadFile = File(...), admin: dict = Depends(require_settings_admin)
):
    """从 xlsx 批量导入维表取值：同表 code 已存在则更新，否则新增。列与校验由字段规格驱动。"""
    table = _dim_table_or_404(table_id)
    spec = _dim_value_field_spec(table["code"])
    spec_by_key = {f["key"]: f for f in spec}
    keys = [f["key"] for f in spec]

    raw = await file.read()
    original_name = file.filename or "dim_import.xlsx"
    username = admin.get("username") or "admin"
    if len(raw) > 10 * 1024 * 1024:
        _write_transfer_meta(
            "import", username, original_name, len(raw), _XLSX_MIME,
            remark=f"维表「{table['name']}」导入：文件过大（上限 10MB）", status="failed",
        )
        raise HTTPException(status_code=400, detail="文件过大（上限 10MB）")

    saved_path, record_id = _write_transfer_record(
        "import", username, original_name, raw, _XLSX_MIME,
        remark=f"维表「{table['name']}」导入", status="",
    )

    def _mark_failed(reason: str) -> None:
        execute(
            "UPDATE transfer_records SET status='failed', remark=? WHERE id=?",
            (f"维表「{table['name']}」导入：{reason}", record_id),
        )

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
    except Exception:
        _mark_failed("无法解析文件")
        raise HTTPException(status_code=400, detail="无法解析文件，请上传有效的 .xlsx 文件")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        _mark_failed("空文件")
        raise HTTPException(status_code=400, detail="空文件")

    header = [str(h).strip().lower() if h is not None else "" for h in header_row]
    for must in ("code", "name"):
        if must not in header:
            _mark_failed(f"缺少 {must} 列")
            raise HTTPException(status_code=400, detail=f"xlsx 缺少 {must} 列")
    # 只识别规格内、且表头里存在的列
    idx = {h: i for i, h in enumerate(header)}
    active_keys = [k for k in keys if k in idx]

    def cell(row, key: str, default=""):
        i = idx.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return default
        return row[i]

    created = 0
    updated = 0
    errors: list[str] = []
    row_errors: dict[int, str] = {}
    seen_codes: set[str] = set()

    for n, row in enumerate(rows_iter, start=2):
        if row is None or not any(c is not None and str(c).strip() for c in row):
            continue
        code = str(cell(row, "code", "")).strip()
        name = str(cell(row, "name", "")).strip()
        if not code or not name:
            reason = "code/name 不能为空"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue
        if code in seen_codes:
            reason = f"文件内编码重复：{code}"
            errors.append(f"第 {n} 行：{reason}")
            row_errors[n] = reason
            continue

        # 按规格逐列解析 + 校验
        values: dict[str, object] = {"code": code, "name": name}
        bad = None
        for k in active_keys:
            f = spec_by_key[k]
            raw_v = cell(row, k, "" if not f.get("is_bool") and not f.get("is_int") else 0)
            if f.get("is_bool"):
                values[k] = _parse_bool(raw_v)
            elif f.get("is_int"):
                try:
                    values[k] = int(float(raw_v if raw_v != "" else f.get("default", 0)))
                except (ValueError, TypeError):
                    values[k] = f.get("default", 0)
            else:
                sval = str(raw_v or "").strip()
                if f.get("required") and not sval:
                    bad = f"{k} 不能为空"
                    break
                if f.get("max_len") and len(sval) > f["max_len"]:
                    bad = f"{k} 长度超限（{f['max_len']}）"
                    break
                if f.get("no_cjk") and _CJK_RE.search(sval):
                    bad = f"{k} 不允许包含中文"
                    break
                values[k] = sval
        if bad:
            errors.append(f"第 {n} 行：{bad}")
            row_errors[n] = bad
            continue

        # 启用校验：模型供应商（ollama 豁免）要启用必须带 api_key
        row_enabled = bool(values.get("enabled", 1))
        enable_bad = _validate_provider_enable(
            table_id, code, str(values.get("api_key", "")), row_enabled
        )
        if enable_bad:
            errors.append(f"第 {n} 行：{enable_bad}")
            row_errors[n] = enable_bad
            continue

        exists = fetch_one(
            "SELECT id FROM dim_values WHERE table_id = ? AND code = ?", (table_id, code)
        )
        col_keys = [k for k in active_keys]
        if exists:
            set_clause = ", ".join(f"{k} = ?" for k in col_keys)
            params = [values[k] for k in col_keys] + [now_ms(), exists["id"]]
            execute(
                f"UPDATE dim_values SET {set_clause}, updated_at = ? WHERE id = ?", params
            )
            updated += 1
        else:
            cols = ["table_id"] + col_keys + ["created_at", "updated_at"]
            placeholders = ", ".join(["?"] * len(cols))
            params = [table_id] + [values[k] for k in col_keys] + [now_ms(), now_ms()]
            execute(f"INSERT INTO dim_values ({', '.join(cols)}) VALUES ({placeholders})", params)
            created += 1
        seen_codes.add(code)

    if row_errors and (created + updated) == 0:
        status = "failed"
    elif row_errors:
        status = "partial"
    else:
        status = "success"

    if row_errors:
        try:
            ann_path, ann_size = _build_annotated_file(raw, row_errors, original_name)
            if os.path.isfile(saved_path):
                os.remove(saved_path)
            execute(
                "UPDATE transfer_records SET file_path=?, file_size=? WHERE id=?",
                (ann_path, ann_size, record_id),
            )
        except Exception:
            pass

    execute("UPDATE transfer_records SET status=? WHERE id=?", (status, record_id))
    return {"ok": True, "created": created, "updated": updated, "errors": errors}


# ============ 导入 / 导出记录（所有管理员可见，可下载源文件） ============

_TRANSFER_DIR = os.path.join(os.path.dirname(settings.db_path) or ".", "transfers")


@router.get("/transfers")
def list_transfers(
    type: str = Query("import", description="import / export"),
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    status: str = "",
    username: str = "",
    sort: str = "",
    order: str = "desc",
):
    if type not in ("import", "export"):
        raise HTTPException(status_code=400, detail="type 仅支持 import/export")
    offset = (page - 1) * page_size
    clauses = ["type = ?"]
    params: list = [type]
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if statuses:
            clauses.append("status IN (" + ",".join("?" for _ in statuses) + ")")
            params.extend(statuses)
    if username:
        usernames = [u.strip() for u in username.split(",") if u.strip()]
        if usernames:
            clauses.append("username IN (" + ",".join("?" for _ in usernames) + ")")
            params.extend(usernames)
    where = "WHERE " + " AND ".join(clauses)
    sort_columns = {
        "id": "id",
        "filename": "filename",
        "username": "username",
        "file_size": "file_size",
        "created_at": "created_at",
    }
    order_by = "created_at DESC, id DESC"
    if sort in sort_columns:
        direction = "ASC" if order.strip().lower() == "asc" else "DESC"
        order_by = f"{sort_columns[sort]} {direction}, id DESC"
    total = fetch_one(f"SELECT COUNT(*) AS n FROM transfer_records {where}", params)["n"]
    rows = fetch_all(
        f"SELECT * FROM transfer_records {where} ORDER BY {order_by} LIMIT ? OFFSET ?",
        (*params, page_size, offset),
    )
    items = []
    for r in rows:
        d = dict(r)
        d["has_file"] = bool(d.get("file_path")) and os.path.isfile(d["file_path"])
        items.append(d)
    return {"items": items, "total": total, "page": page, "pageSize": page_size}


@router.delete("/transfers/{record_id}")
def delete_transfer(record_id: int):
    """删除一条导入/导出记录；导入源文件一并清理（导出本就不存产物）。"""
    row = fetch_one("SELECT * FROM transfer_records WHERE id = ?", (record_id,))
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    file_path = row["file_path"]
    if file_path:
        base = os.path.abspath(_TRANSFER_DIR)
        full = os.path.abspath(file_path)
        if full == base or full.startswith(base + os.sep):
            try:
                if os.path.isfile(full):
                    os.remove(full)
            except OSError:
                pass
    execute("DELETE FROM transfer_records WHERE id = ?", (record_id,))
    return {"ok": True}


@router.get("/admins")
def list_admins():
    """返回所有管理员用户名（供导入/导出记录的操作人筛选下拉框）。"""
    rows = fetch_all(
        "SELECT DISTINCT username FROM users "
        "WHERE role IN ('super_admin','system_admin','model_admin') ORDER BY username"
    )
    return [r["username"] for r in rows]


@router.get("/transfers/{record_id}/download")
def download_transfer(record_id: int):
    row = fetch_one("SELECT * FROM transfer_records WHERE id = ?", (record_id,))
    if not row:
        raise HTTPException(status_code=404, detail="记录不存在")
    file_path = row["file_path"]
    if not file_path:
        raise HTTPException(status_code=404, detail="该记录没有源文件")
    # 防路径穿越：强制限制在 transfers 目录内
    base = os.path.abspath(_TRANSFER_DIR)
    full = os.path.abspath(file_path)
    if not (full == base or full.startswith(base + os.sep)):
        raise HTTPException(status_code=400, detail="非法文件路径")
    if not os.path.isfile(full):
        raise HTTPException(status_code=404, detail="源文件不存在或已被清理")
    media_type = row["mime_type"] or "application/octet-stream"
    return FileResponse(full, filename=row["filename"], media_type=media_type)