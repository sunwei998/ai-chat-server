"""端到端验证：维表字段配置化（dim_table_fields）。

用法（后端已启动在 127.0.0.1:8000）：
    .venv/bin/python scripts/e2e_dim_fields_config.py

覆盖：
1. 新建维表自动写入默认字段配置；读取接口返回 persisted=true
2. 关闭某列后，导出与模板都不再包含该列
3. 改列头标签后，导出用新标签，且「旧标签表头」的历史文件仍能导入
4. 关掉的列在导入时不更新，库里的历史值保留
5. code / name 强制启用且必填，忽略请求里的相反设置
6. 非法 field_key 与非法 field_type 被拒
7. 清空配置后回退代码内置模板，导出导入照常工作
"""

import io
import json
import sys
import urllib.error
import urllib.request
import uuid

import openpyxl

BASE = "http://127.0.0.1:8000"
USERNAME, PASSWORD = "admin", "admin123"

PASS, FAIL = [], []


def check(name, ok, detail=""):
    (PASS if ok else FAIL).append(name)
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def req(method, path, lang="zh-CN", token=None, data=None, files=None, raw=False):
    headers = {"Accept-Language": lang}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None
    if data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
    elif files is not None:
        boundary = "----dim" + uuid.uuid4().hex
        buf = io.BytesIO()
        for field, (fname, content) in files.items():
            buf.write(f"--{boundary}\r\n".encode())
            buf.write(
                f'Content-Disposition: form-data; name="{field}"; filename="{fname}"\r\n'.encode()
            )
            buf.write(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n")
            buf.write(content)
            buf.write(b"\r\n")
        buf.write(f"--{boundary}--\r\n".encode())
        body = buf.getvalue()
        headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    r = urllib.request.Request(BASE + path, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r) as resp:
            payload = resp.read()
            return resp.status, (payload if raw else json.loads(payload))
    except urllib.error.HTTPError as e:
        payload = e.read()
        try:
            return e.code, json.loads(payload)
        except Exception:
            return e.code, payload.decode("utf-8", "replace")


def rows_of(xlsx_bytes, max_row=3):
    ws = openpyxl.load_workbook(io.BytesIO(xlsx_bytes)).active
    return [list(r) for r in ws.iter_rows(min_row=1, max_row=max_row, values_only=True)]


def build_xlsx(header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    status, res = req("POST", "/api/auth/login", data={"username": USERNAME, "password": PASSWORD})
    if status != 200 or not isinstance(res, dict) or "token" not in res:
        print("登录失败：", status, res)
        return 1
    token = res["token"]
    print(f"登录成功（{USERNAME}）\n")

    code = "zzfieldcfg"
    st, tables = req("GET", "/api/admin/dim-tables", "zh-CN", token)
    if isinstance(tables, list):
        for t in tables:
            if t.get("code") == code:
                req("DELETE", f"/api/admin/dim-tables/{t['id']}", "zh-CN", token)
    st, tbl = req("POST", "/api/admin/dim-tables", "zh-CN", token,
                  data={"code": code, "name": "字段配置测试表", "description": "e2e"})
    if st != 200 or not isinstance(tbl, dict) or "id" not in tbl:
        print("创建临时维表失败：", st, tbl)
        return 1
    tid = tbl["id"]
    print(f"临时维表 id={tid}\n")

    # ---------- 1. 默认配置 ----------
    print("1) 新建维表自动写入默认字段配置")
    st, flds = req("GET", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token)
    keys = [f["field_key"] for f in (flds or {}).get("items", [])]
    check("GET fields 返回 200", st == 200, f"status={st}")
    check("persisted=true（已落库）", (flds or {}).get("persisted") is True, str((flds or {}).get("persisted")))
    check("默认 5 列且顺序正确", keys == ["code", "name", "sort_order", "enabled", "remark"], str(keys))

    st, zh_exp = req("GET", f"/api/admin/dim-tables/{tid}/export", "zh-CN", token, raw=True)
    check("导出表头 = 默认中文标签",
          rows_of(zh_exp, 1)[0] == ["编码", "名称", "排序", "启用", "备注"],
          str(rows_of(zh_exp, 1)[0]))

    # ---------- 2. 关闭列 ----------
    print("\n2) 关闭某列后，导出与模板都不再包含该列")
    items = (flds or {}).get("items", [])
    for f in items:
        if f["field_key"] == "remark":
            f["enabled"] = False
    st, saved = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token, data={"items": items})
    check("PUT fields 保存成功", st == 200 and (saved or {}).get("ok") is True, f"status={st} {saved}")

    st, exp2 = req("GET", f"/api/admin/dim-tables/{tid}/export", "zh-CN", token, raw=True)
    check("导出已移除该列", rows_of(exp2, 1)[0] == ["编码", "名称", "排序", "启用"],
          str(rows_of(exp2, 1)[0]))
    st, tpl2 = req("GET", f"/api/admin/dim-tables/{tid}/template", "zh-CN", token, raw=True)
    check("模板已移除该列", rows_of(tpl2, 1)[0] == ["编码", "名称", "排序", "启用"],
          str(rows_of(tpl2, 1)[0]))
    check("模板示例行长度跟着变", len(rows_of(tpl2, 2)[1]) == 4, str(rows_of(tpl2, 2)[1]))

    # ---------- 3. 改标签 ----------
    print("\n3) 改列头标签后，导出用新标签、旧标签历史文件仍能导入")
    st, flds2 = req("GET", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token)
    items2 = (flds2 or {}).get("items", [])
    for f in items2:
        if f["field_key"] == "name":
            f["label_zh"], f["label_en"] = "显示名", "Display Name"
    st, _ = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token, data={"items": items2})
    st, exp3 = req("GET", f"/api/admin/dim-tables/{tid}/export", "zh-CN", token, raw=True)
    check("中文导出用新标签", rows_of(exp3, 1)[0][1] == "显示名", str(rows_of(exp3, 1)[0]))
    st, exp3en = req("GET", f"/api/admin/dim-tables/{tid}/export", "en-US", token, raw=True)
    check("英文导出用新标签", rows_of(exp3en, 1)[0][1] == "Display Name", str(rows_of(exp3en, 1)[0]))

    f_new = build_xlsx(["编码", "显示名", "排序", "启用"], [["alpha", "阿尔法", 1, "是"]])
    st, r1 = req("POST", f"/api/admin/dim-tables/{tid}/import", "zh-CN", token,
                 files={"file": ("new.xlsx", f_new)})
    check("新标签表头可导入", st == 200 and (r1 or {}).get("created", 0) >= 1, f"status={st} {r1}")

    f_old = build_xlsx(["编码", "名称", "排序", "启用"], [["beta", "贝塔", 2, "是"]])
    st, r2 = req("POST", f"/api/admin/dim-tables/{tid}/import", "zh-CN", token,
                 files={"file": ("old.xlsx", f_old)})
    check("旧标签（改名前的历史文件）仍可导入",
          st == 200 and (r2 or {}).get("created", 0) >= 1, f"status={st} {r2}")

    # ---------- 4. 关闭的列导入时保留原值 ----------
    print("\n4) 关掉的列在导入时不更新，库里历史值保留")
    st, vals = req("GET", f"/api/admin/dim-tables/{tid}/values?page_size=50", "zh-CN", token)
    target_id = None
    for v in (vals or {}).get("items", []):
        if v.get("code") == "alpha":
            target_id = v["id"]
    st, _ = req("PUT", f"/api/admin/dim-tables/{tid}/values/{target_id}", "zh-CN", token,
                data={"remark": "原始备注"})
    f_upd = build_xlsx(["编码", "显示名", "排序", "启用"], [["alpha", "阿尔法改", 9, "否"]])
    st, r3 = req("POST", f"/api/admin/dim-tables/{tid}/import", "zh-CN", token,
                 files={"file": ("upd.xlsx", f_upd)})
    st, vals2 = req("GET", f"/api/admin/dim-tables/{tid}/values?page_size=50", "zh-CN", token)
    after = next((v for v in (vals2 or {}).get("items", []) if v.get("code") == "alpha"), {})
    check("启用列照常更新", after.get("name") == "阿尔法改", str(after.get("name")))
    check("关闭列保留原值（remark 未被清空）", after.get("remark") == "原始备注",
          str(after.get("remark")))

    # ---------- 5. code/name 强制启用 ----------
    print("\n5) code / name 强制启用且必填")
    st, flds3 = req("GET", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token)
    items3 = (flds3 or {}).get("items", [])
    for f in items3:
        if f["field_key"] in ("code", "name"):
            f["enabled"] = False
            f["required"] = False
    st, _ = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token, data={"items": items3})
    st, flds4 = req("GET", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token)
    core = {f["field_key"]: f for f in (flds4 or {}).get("items", [])}
    check("code 仍为 enabled+required",
          core.get("code", {}).get("enabled") == 1 and core.get("code", {}).get("required") == 1,
          str(core.get("code")))
    check("name 仍为 enabled+required",
          core.get("name", {}).get("enabled") == 1 and core.get("name", {}).get("required") == 1,
          str(core.get("name")))
    st, exp5 = req("GET", f"/api/admin/dim-tables/{tid}/export", "zh-CN", token, raw=True)
    check("code/name 仍在导出表头里", "编码" in rows_of(exp5, 1)[0] and "显示名" in rows_of(exp5, 1)[0],
          str(rows_of(exp5, 1)[0]))

    # ---------- 6. 非法输入被拒 ----------
    print("\n6) 非法字段配置被拒")
    st, bad1 = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token,
                   data={"items": [{"field_key": "drop_table", "label_zh": "x"}]})
    check("非物理列 field_key 被拒 422", st == 422, f"status={st} {bad1}")
    st, bad2 = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token,
                   data={"items": [{"field_key": "remark", "field_type": "blob"}]})
    check("非法 field_type 被拒 422", st == 422, f"status={st} {bad2}")

    # ---------- 7. 清空配置后回退模板 ----------
    print("\n7) 清空配置后回退代码内置模板")
    st, flds5 = req("GET", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token)
    all_off = (flds5 or {}).get("items", [])
    for f in all_off:
        f["enabled"] = False
    st, _ = req("PUT", f"/api/admin/dim-tables/{tid}/fields", "zh-CN", token, data={"items": all_off})
    st, exp7 = req("GET", f"/api/admin/dim-tables/{tid}/export", "zh-CN", token, raw=True)
    check("全部关闭时仍回退出 code/name（核心列兜底）",
          rows_of(exp7, 1)[0] == ["编码", "显示名"], str(rows_of(exp7, 1)[0]))

    # ---------- 清理 ----------
    st, _ = req("DELETE", f"/api/admin/dim-tables/{tid}", "zh-CN", token)
    check("\n清理临时维表", st == 200, f"status={st}")

    print(f"\n========== 通过 {len(PASS)} / 失败 {len(FAIL)} ==========")
    if FAIL:
        print("失败项：")
        for f in FAIL:
            print("  -", f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
