"""端到端验证：导入导出表头国际化 + 中英文表头交叉导入兼容。

用法（后端已启动在 127.0.0.1:8000）：
    .venv/bin/python scripts/e2e_i18n_import_export.py

覆盖：
1. 中文环境导出/模板 → 中文表头；英文环境 → 英文表头（models / settings / dim / users）
2. 中文表头文件在英文环境下导入成功；英文表头文件在中文环境下导入成功
3. 导入错误提示语言跟随请求语言
"""

import io
import json
import sys
import urllib.error
import urllib.request
import uuid

import openpyxl

BASE = "http://127.0.0.1:8000"
USERNAME, PASSWORD = "admin", "admin123"

PASS, FAIL = [], []


def check(name, ok, detail=""):
    (PASS if ok else FAIL).append(name)
    print(f"  [{'PASS' if ok else 'FAIL'}] {name}" + (f" — {detail}" if detail else ""))


def req(method, path, lang="zh-CN", token=None, data=None, files=None, raw=False):
    url = BASE + path
    headers = {"Accept-Language": lang}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    body = None
    if data is not None:
        body = json.dumps(data).encode()
        headers["Content-Type"] = "application/json"
    elif files is not None:
        boundary = "----i18n" + uuid.uuid4().hex
        buf = io.BytesIO()
        for field, (fname, content) in files.items():
            buf.write(f"--{boundary}\r\n".encode())
            buf.write(
                f'Content-Disposition: form-data; name="{field}"; filename="{fname}"\r\n'.encode()
            )
            buf.write(b"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n")
            buf.write(content)
            buf.write(b"\r\n")
        buf.write(f"--{boundary}--\r\n".encode())
        body = buf.getvalue()
        headers["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    r = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(r) as resp:
            payload = resp.read()
            return resp.status, (payload if raw else json.loads(payload))
    except urllib.error.HTTPError as e:
        payload = e.read()
        try:
            return e.code, json.loads(payload)
        except Exception:
            return e.code, payload.decode("utf-8", "replace")


def sheet_rows(xlsx_bytes, max_row=3):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    return [list(r) for r in ws.iter_rows(min_row=1, max_row=max_row, values_only=True)]


def build_xlsx(header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    status, res = req("POST", "/api/auth/login", data={"username": USERNAME, "password": PASSWORD})
    if status != 200 or not isinstance(res, dict) or "token" not in res:
        print("登录失败：", status, res)
        return 1
    token = res["token"]
    print(f"登录成功（{USERNAME}）\n")

    # ---------- 1. 模板表头本地化 ----------
    print("1) 模板表头按语言输出")
    st, zh_tpl = req("GET", "/api/admin/models/template", "zh-CN", token, raw=True)
    check("models/template zh-CN 200", st == 200, f"status={st}")
    zh_hdr = sheet_rows(zh_tpl, 1)[0]
    st, en_tpl = req("GET", "/api/admin/models/template", "en-US", token, raw=True)
    en_hdr = sheet_rows(en_tpl, 1)[0]
    check(
        "models 模板中文表头",
        zh_hdr == ["模型标识", "名称", "英文名称", "提供商", "免费", "视觉", "联网", "启用", "排序", "默认"],
        str(zh_hdr),
    )
    check(
        "models 模板英文表头",
        en_hdr == ["Model Key", "Name", "English Name", "Provider", "Free", "Vision",
                   "Web Search", "Enabled", "Sort", "Default"],
        str(en_hdr),
    )

    st, zh_set_tpl = req("GET", "/api/admin/settings/template", "zh-CN", token, raw=True)
    st, en_set_tpl = req("GET", "/api/admin/settings/template", "en-US", token, raw=True)
    check("settings 模板中文表头", sheet_rows(zh_set_tpl, 1)[0] == ["键", "值", "备注", "启用"],
          str(sheet_rows(zh_set_tpl, 1)[0]))
    check("settings 模板英文表头", sheet_rows(en_set_tpl, 1)[0] == ["Key", "Value", "Remark", "Enabled"],
          str(sheet_rows(en_set_tpl, 1)[0]))

    # ---------- 2. 导出表头本地化 ----------
    print("\n2) 导出表头按语言输出")
    st, zh_exp = req("GET", "/api/admin/models/export?scope=all", "zh-CN", token, raw=True)
    st, en_exp = req("GET", "/api/admin/models/export?scope=all", "en-US", token, raw=True)
    check("models 导出中文表头", sheet_rows(zh_exp, 1)[0] == list(zh_hdr), str(sheet_rows(zh_exp, 1)[0]))
    check("models 导出英文表头", sheet_rows(en_exp, 1)[0] == list(en_hdr), str(sheet_rows(en_exp, 1)[0]))

    st, zh_users = req("GET", "/api/admin/users/export", "zh-CN", token, raw=True)
    st, en_users = req("GET", "/api/admin/users/export", "en-US", token, raw=True)
    check("users 导出中文表头", "用户名" in sheet_rows(zh_users, 1)[0], str(sheet_rows(zh_users, 1)[0]))
    check("users 导出英文表头", "Username" in sheet_rows(en_users, 1)[0], str(sheet_rows(en_users, 1)[0]))

    st, zh_set_exp = req("GET", "/api/admin/settings/export", "zh-CN", token, raw=True)
    st, en_set_exp = req("GET", "/api/admin/settings/export", "en-US", token, raw=True)
    check("settings 导出中文表头", sheet_rows(zh_set_exp, 1)[0] == ["键", "值", "备注", "启用"],
          str(sheet_rows(zh_set_exp, 1)[0]))
    check("settings 导出英文表头", sheet_rows(en_set_exp, 1)[0] == ["Key", "Value", "Remark", "Enabled"],
          str(sheet_rows(en_set_exp, 1)[0]))

    # ---------- 3. 维表（建临时表） ----------
    print("\n3) 维表导入导出（临时维表）")
    code = "zz_i18n_t"
    st, tables = req("GET", "/api/admin/dim-tables", "zh-CN", token)
    table_id = None
    if isinstance(tables, list):
        for t in tables:
            if t.get("code") == code:
                table_id = t["id"]
                req("DELETE", f"/api/admin/dim-tables/{table_id}", "zh-CN", token)
    st, tbl = req("POST", "/api/admin/dim-tables", "zh-CN", token,
                  data={"code": code, "name": "国际化测试表", "description": "e2e"})
    if st == 200 and isinstance(tbl, dict):
        table_id = tbl["id"]
    check("创建临时维表", table_id is not None, f"status={st} {tbl}")

    dim_ok = table_id is not None
    if dim_ok:
        st, zh_dim_tpl = req("GET", f"/api/admin/dim-tables/{table_id}/template", "zh-CN", token, raw=True)
        st, en_dim_tpl = req("GET", f"/api/admin/dim-tables/{table_id}/template", "en-US", token, raw=True)
        check("dim 模板中文表头",
              sheet_rows(zh_dim_tpl, 1)[0] == ["编码", "名称", "排序", "启用", "备注"],
              str(sheet_rows(zh_dim_tpl, 1)[0]))
        check("dim 模板英文表头",
              sheet_rows(en_dim_tpl, 1)[0] == ["Code", "Name", "Sort", "Enabled", "Remark"],
              str(sheet_rows(en_dim_tpl, 1)[0]))

    # ---------- 4. 交叉导入 ----------
    print("\n4) 中英文表头交叉导入（中文模式导入英文模板，反之亦然）")
    import random, string
    suffix = "".join(random.choice(string.ascii_lowercase) for _ in range(6))
    PROV = "ollama"  # 维表 model_provider 中实际启用且免 api_key 的值

    # 4a. models：中文表头文件 → 英文环境导入
    zh_rows = [
        [f"zz_i18n_{suffix}_a", f"中文表头模型A_{suffix}", "Model A", PROV, "否", "否", "是", "是", 900, "否"],
    ]
    f_zh = build_xlsx(list(zh_hdr), zh_rows)
    st, r1 = req("POST", "/api/admin/models/import", "en-US", token, files={"file": ("zh.xlsx", f_zh)})
    check("中文表头文件 → 英文环境导入 models",
          st == 200 and isinstance(r1, dict) and r1.get("created", 0) >= 1, f"status={st} {r1}")

    # 4b. models：英文表头文件 → 中文环境导入
    en_rows = [
        [f"zz_i18n_{suffix}_b", f"英文表头模型B_{suffix}", "Model B", PROV, "no", "no", "yes", "yes", 901, "no"],
    ]
    f_en = build_xlsx(list(en_hdr), en_rows)
    st, r2 = req("POST", "/api/admin/models/import", "zh-CN", token, files={"file": ("en.xlsx", f_en)})
    check("英文表头文件 → 中文环境导入 models",
          st == 200 and isinstance(r2, dict) and r2.get("created", 0) >= 1, f"status={st} {r2}")

    # 4c. 旧版英文字段名表头仍可导入（向后兼容）
    raw_hdr = ["model_key", "name", "name_en", "provider", "free", "vision",
               "supports_search", "enabled", "sort_order", "is_default"]
    f_raw = build_xlsx(raw_hdr, [[f"zz_i18n_{suffix}_c", f"原始表头模型C_{suffix}", "Model C",
                                  PROV, 0, 0, 1, 1, 902, 0]])
    st, r3 = req("POST", "/api/admin/models/import", "en-US", token, files={"file": ("raw.xlsx", f_raw)})
    check("原始英文字段名表头仍可导入",
          st == 200 and isinstance(r3, dict) and r3.get("created", 0) >= 1, f"status={st} {r3}")

    # 4d. settings 交叉导入
    f_set_zh = build_xlsx(["键", "值", "备注", "启用"],
                          [[f"zztest_{suffix}", "12345", "中文表头导入", "是"]])
    st, r4 = req("POST", "/api/admin/settings/import", "en-US", token, files={"file": ("s.xlsx", f_set_zh)})
    check("中文表头文件 → 英文环境导入 settings",
          st == 200 and isinstance(r4, dict) and r4.get("created", 0) >= 1, f"status={st} {r4}")

    f_set_en = build_xlsx(["Key", "Value", "Remark", "Enabled"],
                          [[f"zztest_{suffix}en", "54321", "English header import", "yes"]])
    st, r5 = req("POST", "/api/admin/settings/import", "zh-CN", token, files={"file": ("s.xlsx", f_set_en)})
    check("英文表头文件 → 中文环境导入 settings",
          st == 200 and isinstance(r5, dict) and r5.get("created", 0) >= 1, f"status={st} {r5}")

    # 4e. dim 交叉导入
    if dim_ok:
        f_dim_zh = build_xlsx(["编码", "名称", "排序", "启用", "备注"],
                              [[f"zh_{suffix}", "中文表头取值", 1, "是", "备注"]])
        st, r6 = req("POST", f"/api/admin/dim-tables/{table_id}/import", "en-US", token,
                     files={"file": ("d.xlsx", f_dim_zh)})
        check("中文表头文件 → 英文环境导入 dim",
              st == 200 and isinstance(r6, dict) and r6.get("created", 0) >= 1, f"status={st} {r6}")

        f_dim_en = build_xlsx(["Code", "Name", "Sort", "Enabled", "Remark"],
                              [[f"en_{suffix}", "English Header Value", 2, "yes", "remark"]])
        st, r7 = req("POST", f"/api/admin/dim-tables/{table_id}/import", "zh-CN", token,
                     files={"file": ("d.xlsx", f_dim_en)})
        check("英文表头文件 → 中文环境导入 dim",
              st == 200 and isinstance(r7, dict) and r7.get("created", 0) >= 1, f"status={st} {r7}")

    # ---------- 5. 错误提示语言 ----------
    print("\n5) 导入错误提示跟随语言")
    bad_zh = build_xlsx(list(zh_hdr), [["", "缺主键", "", "openai", "否", "否", "是", "是", 1, "否"]])
    st, e1 = req("POST", "/api/admin/models/import", "zh-CN", token, files={"file": ("b.xlsx", bad_zh)})
    check("中文环境错误为中文",
          isinstance(e1, dict) and any("第 2 行" in x for x in e1.get("errors", [])), str(e1))
    st, e2 = req("POST", "/api/admin/models/import", "en-US", token, files={"file": ("b.xlsx", bad_zh)})
    check("英文环境错误为英文",
          isinstance(e2, dict) and any("Row 2" in x for x in e2.get("errors", [])), str(e2))

    # 缺列提示
    f_nocol = build_xlsx(["名称", "提供商"], [["x", PROV]])
    st, e3 = req("POST", "/api/admin/models/import", "en-US", token, files={"file": ("n.xlsx", f_nocol)})
    check("缺列提示为英文", isinstance(e3, dict) and "missing" in str(e3.get("detail", "")), str(e3))
    st, e4 = req("POST", "/api/admin/models/import", "zh-CN", token, files={"file": ("n.xlsx", f_nocol)})
    check("缺列提示为中文", isinstance(e4, dict) and "缺少" in str(e4.get("detail", "")), str(e4))

    # ---------- 清理 ----------
    print("\n6) 清理测试数据")
    st, models = req("GET", "/api/admin/models?search=zz_i18n&page_size=100", "zh-CN", token)
    removed = 0
    if isinstance(models, dict):
        for m in models.get("items", []):
            if f"zz_i18n_{suffix}" in (m.get("model_key") or ""):
                rs, _ = req("DELETE", f"/api/admin/models/{m['id']}", "zh-CN", token)
                removed += 1 if rs == 200 else 0
    check("清理测试模型", removed >= 3, f"removed={removed}")

    st, settings = req("GET", f"/api/admin/settings?search=zztest&page_size=100", "zh-CN", token)
    sremoved = 0
    if isinstance(settings, dict):
        for s in settings.get("items", []):
            if "zztest_" in (s.get("key") or ""):
                # 启用的配置项不能直接删，先禁用
                req("PATCH", f"/api/admin/settings/{s['key']}", "zh-CN", token,
                    data={"enabled": False})
                rs, _ = req("DELETE", f"/api/admin/settings/{s['key']}", "zh-CN", token)
                sremoved += 1 if rs == 200 else 0
    check("清理测试配置项", sremoved >= 2, f"removed={sremoved}")

    if dim_ok:
        st, _ = req("DELETE", f"/api/admin/dim-tables/{table_id}", "zh-CN", token)
        check("清理临时维表", st == 200, f"status={st}")

    print(f"\n========== 通过 {len(PASS)} / 失败 {len(FAIL)} ==========")
    if FAIL:
        print("失败项：")
        for f in FAIL:
            print("  -", f)
    return 1 if FAIL else 0


if __name__ == "__main__":
    sys.exit(main())
